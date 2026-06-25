# -*- coding: utf-8 -*-
from __future__ import absolute_import, unicode_literals

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseForbidden, HttpResponse, JsonResponse
from django.utils.timezone import is_naive, localtime

from rest_framework import status
from rest_framework import viewsets, mixins, permissions
from django_filters.views import FilterView
from django_tables2 import RequestConfig, SingleTableView
from django_tables2.export.views import ExportMixin
from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import load_workbook
from django.core.files.base import ContentFile
import csv
import io
import re

from student_registration.adolescent.models import Adolescent
from student_registration.students.models import Nationality, IDType
from student_registration.students.utils import generate_bulk_unique_id
from student_registration.clm.models import Disability, EducationalLevel
from student_registration.locations.models import Location
from student_registration.youth.models import Registration, YOUTH_EDUCATION_STATUS




from .exporter import export_full_data
from .models import Notification, Exporter, AdolescentUpload, ExportHistory
from .serializers import NotificationSerializer, ExporterSerializer
from .filters import ExporterFilter
from .tables import BootstrapTable, ExporterTable
from collections import OrderedDict


def generate_child_unique_id(request):
    from student_registration.backends.threads import generate_child_unique_id

    generate_child_unique_id()
    return HttpResponse("records saved successfully")


def generate_all_child_unique_id(request):
    from student_registration.backends.threads import generate_all_child_unique_id

    generate_all_child_unique_id()
    return HttpResponse("records saved successfully")


def generate_child_cash_programme(request):
    from student_registration.backends.threads import generate_child_programmes

    generate_child_programmes()
    return HttpResponse("records saved successfully")


def generate_student_unique_id(request):
    from student_registration.backends.threads import generate_student_unique_id

    generate_student_unique_id()
    return HttpResponse("records saved successfully")



def generate_adolescent_unique_id(request):
    from student_registration.backends.threads import generate_adolescent_unique_id

    generate_adolescent_unique_id()
    return HttpResponse("records saved successfully")


def generate_all_teacher_unique_id(request):
    from student_registration.backends.threads import generate_all_teacher_unique_id

    generate_all_teacher_unique_id()
    return HttpResponse("records saved successfully")



class NotificationViewSet(mixins.UpdateModelMixin,
                          viewsets.GenericViewSet):

    model = Notification
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = (permissions.IsAuthenticated,)

    # def update(self, request, *args, **kwargs):
    #     if 'pk' not in kwargs:
    #         return super(NotificationViewSet, self).update(request)
    #     instance = self.model.objects.get(id=kwargs['pk'])
    #     print(request)
    #     instance.status = True
    #     instance.save()
    #     return JsonResponse({'status': status.HTTP_200_OK, 'data': instance.id})


class ExporterListView(LoginRequiredMixin,
                       FilterView,
                       ExportMixin,
                       SingleTableView,
                       RequestConfig):

    table_class = ExporterTable
    model = Exporter
    template_name = 'backends/files.html'
    table = BootstrapTable(Exporter.objects.all(), order_by='-id')

    filterset_class = ExporterFilter

    def get_queryset(self):
        return Exporter.objects.filter(exported_by=self.request.user)


class ExporterViewSet(LoginRequiredMixin,
                      mixins.ListModelMixin,
                      viewsets.GenericViewSet,):

    model = Exporter
    queryset = Exporter.objects.all()
    serializer_class = ExporterSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def handle_no_permission(self):
        return HttpResponseForbidden()

    def list(self, request, *args, **kwargs):
        if self.request.GET.get('report', None):
            #  todo raise a exception if the partner
            data = {
                'report': self.request.GET.get('report'),
                'user': self.request.user.id,
                'partner': self.request.user.partner_id
            }
            export_full_data(data)
        return JsonResponse({'status': status.HTTP_200_OK})


@login_required
def export_history_list(request):
    exports = ExportHistory.objects.filter(created_by=request.user)
    export_id = request.GET.get('export_id')
    if export_id:
        exports = exports.filter(id=export_id)
    exports = exports.order_by('-created')[:5]

    data = []
    for export in exports:
        created = export.created if is_naive(export.created) else localtime(export.created)
        timestamp = created.strftime('%Y-%m-%d %H:%M')
        data.append({
            'id': export.id,
            'url': export.file_url or '#',
            'text': f'MSCC export {timestamp}',
            'status': export.status,
        })
    return JsonResponse({'exports': data})


class AdolescentUploadView(LoginRequiredMixin, View):
    template_name = 'backends/adolescent_upload.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return render(request, self.template_name, {'error': 'No file selected'})

        upload = AdolescentUpload.objects.create(file=excel_file, uploaded_by=request.user)
        return redirect('backends:adolescent_upload_confirm', pk=upload.pk)


class AdolescentUploadConfirmView(LoginRequiredMixin, View):
    template_name = 'backends/adolescent_upload_confirm.html'
    result_template = 'backends/adolescent_import_result.html'

    mapping = {
        'Adolescent First Name': 'first_name',
        'Adolescent Father Name': 'father_name',
        'Adolescent Last Name': 'last_name',
        'Adolescent Birthday Year': 'birthday_year',
        'Adolescent Birthday Month': 'birthday_month',
        'Adolescent Birthday Day': 'birthday_day',
        'Gender': 'gender',
        'Adolescent Mother Fullname': 'mother_fullname',
        'Adolescent Nationality': 'nationality',
        'Adolescent Nationality Other': 'nationality_other',
        'Governorate': 'governorate',
        'District': 'district',
        'Cadaster': 'cadaster',
        'Adolescent Address': 'address',
        'Special Need': 'disability',
        'Youth Educational Level When Registering' : 'education_status',
        'Dropout Date': 'dropout_date',
        'Father Educational Level': 'father_educational_level',
        'Mother Educational Level': 'mother_educational_level',
        'First Phone Number': 'first_phone_number',
        'Second Phone Number': 'second_phone_number',
        'Main Caregiver': 'main_caregiver',
        'Main Caregiver Other': 'main_caregiver_other',
        'Caregiver First Name': 'caregiver_first_name',
        'Caregiver Middle Name': 'caregiver_middle_name',
        'Caregiver Last Name': 'caregiver_last_name',
        'Main Caregiver Nationality Name': 'main_caregiver_nationality',
        'Main Caregiver Nationality Other': 'main_caregiver_nationality_other',
        'ID Type': 'id_type',
        'UNHCR Case Number': 'case_number',
        'Cargiver Individual ID': 'parent_individual_case_number',
        'Individual ID of the youth': 'individual_case_number',
        'UNHCR Barcode number (Shifra number)': 'recorded_number',
        'unrwa_number': 'unrwa_number',
        'Syrian National ID number of the Cargiver': 'parent_syrian_national_number',
        'Syrian National ID number of the youth': 'syrian_national_number',
        'Palestinian ID number of the Cargiver': 'parent_sop_national_number',
        'Palestinian ID number of the youth': 'sop_national_number',
        'Lebanese ID number of the Cargiver': 'parent_national_number',
        'Lebanese ID number of the youth': 'national_number',
        'ID number of the Cargiver': 'parent_other_number',
        'ID number of the youth': 'other_number',
    }

    def _build_error_entry(self, row_values, row_number, error_message):
        ordered = OrderedDict()
        ordered['error'] = error_message
        ordered['row'] = row_number
        for field in self.mapping.values():
            ordered[field] = row_values.get(field, '')
        return ordered

    def _is_unicef_id_generation_error(self, generated_value):
        if not generated_value:
            return True

        generated_value = str(generated_value).strip()
        error_terms = ('error', 'invalid', 'unable', 'failed', 'failure', 'translate')
        return any(term in generated_value.lower() for term in error_terms)

    def _unicef_id_generation_error_message(self, generated_value):
        message = "Unable to generate UNICEF ID"
        if generated_value:
            message += ": {0}".format(str(generated_value).strip())

        return message + "."

    phone_number_regex = re.compile(r'^(((03|70|71|76|78|79|81|86)-\d{6})|(963 \d{2} \d{3} \d{4}))$')

    mandatory_fields = [
        'first_name',
        'father_name',
        'last_name',
        'birthday_year',
        'birthday_month',
        'birthday_day',
        'gender',
        'mother_fullname',
        'nationality',
        'governorate',
        'district',
        'cadaster',
        'disability',
        'first_phone_number',
    ]

    def get(self, request, pk, *args, **kwargs):
        try:
            upload = get_object_or_404(AdolescentUpload, pk=pk, uploaded_by=request.user)
            data = self.parse_file(upload.file)

            if isinstance(data, dict) and data.get("error"):
                messages.error(request, data["error"], extra_tags='import')
                return redirect("backends:adolescent_upload")

            preview = data[:5]
            return render(request, self.template_name, {
                'upload': upload,
                'count': len(data),
                'preview': preview,
            })
        except Exception as e:
            messages.error(request, "An unexpected error occurred while processing the file", extra_tags='import')
            return redirect("backends:adolescent_upload")

    def post(self, request, pk, *args, **kwargs):
        try:
            upload = get_object_or_404(AdolescentUpload, pk=pk, uploaded_by=request.user)
            data = self.parse_file(upload.file)

            if isinstance(data, dict) and data.get("error"):
                messages.error(request, data["error"], extra_tags='import')
                return redirect("backends:adolescent_upload")

            imported, not_imported = self.import_data(data, upload, request)
            upload.processed = True
            upload.save()
            return render(request, self.result_template, {
                'imported': imported,
                'failed': len(not_imported),
                'not_imported': not_imported,
                'upload': upload,
            })
        except Exception as e:
            messages.error(request, "An unexpected error occurred while importing the file", extra_tags='import')
            return redirect("backends:adolescent_upload")

    def parse_file(self, uploaded_file):
        from openpyxl import load_workbook
        from collections import OrderedDict

        try:
            wb = load_workbook(filename=uploaded_file, read_only=True)

            if not wb.sheetnames:
                return {"error": "The uploaded Excel file has no sheets. Please check the file content."}

            if 'Registrations' not in wb.sheetnames:
                return {"error": "Sheet 'Registrations' not found in Excel file."}

            ws = wb['Registrations']

            try:
                header_row = next(ws.iter_rows(min_row=1, max_row=1))
                headers = [cell.value for cell in header_row]
            except Exception:
                return {"error": "Failed to read header row in 'Registrations' sheet."}

            rows = []
            skipped_rows = []

            for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    values = [cell.value for cell in row]
                    if not any(values):
                        continue
                    row_data = OrderedDict(
                        (self.mapping.get(headers[i]), values[i])
                        for i in range(len(headers))
                        if headers[i] in self.mapping
                    )
                    row_data['row'] = row_index
                    rows.append(row_data)
                except Exception:
                    skipped_rows.append(row_index)
                    continue

            if skipped_rows:
                return {
                    "error": "Failed to parse rows: {}. Please check the file format.".format(
                        ", ".join(map(str, skipped_rows))
                    ),
                    "skipped": skipped_rows
                }

            return rows

        except Exception:
            return {"error": "Unexpected error while processing Excel file"}

    def import_data(self, data, upload, request):
        import datetime

        not_imported = []
        imported = 0
        validated_rows = []

        for row_data in data:
            values = dict(row_data)
            row_number = values.pop('row', None)
            if row_number is None:
                row_number = len(validated_rows) + len(not_imported) + 2

            # ---- Missing mandatory fields
            missing = [f for f in self.mandatory_fields if not values.get(f)]
            if missing:
                not_imported.append(
                    self._build_error_entry(
                        values,
                        row_number,
                        'Missing fields: ' + ', '.join(missing)
                    )
                )
                continue

            invalid_fields = []

            first_phone_number = str(values.get('first_phone_number') or '').strip()
            second_phone_number = str(values.get('second_phone_number') or '').strip()
            values['first_phone_number'] = first_phone_number
            values['second_phone_number'] = second_phone_number

            if not self.phone_number_regex.match(first_phone_number):
                invalid_fields.append("invalid first phone number")
            if second_phone_number and not self.phone_number_regex.match(second_phone_number):
                invalid_fields.append("invalid second phone number")

            # ---- Normalize gender (and validate)
            raw_gender = (values.get('gender') or '').strip()
            gender_norm = raw_gender.title() if raw_gender else ''
            if gender_norm not in ('Male', 'Female'):
                invalid_fields.append("gender ({0})".format(raw_gender or 'None'))

            raw_main_caregiver = (values.get('main_caregiver') or '').strip()
            main_caregiver_norm = raw_main_caregiver[:1].upper() + raw_main_caregiver[1:] if raw_main_caregiver else ''
            values['main_caregiver'] = main_caregiver_norm

            nationality = Nationality.objects.filter(name_en=values.get('nationality')).first()

            gov_name = (values.get('governorate') or '').strip()
            dist_name = (values.get('district') or '').strip()
            cad_name = (values.get('cadaster') or '').strip()

            gov = Location.objects.filter(name_en=gov_name, type_id=1).first()

            district_qs = Location.objects.none()
            dist = None
            if dist_name:
                district_qs = Location.objects.filter(name_en=dist_name, type_id=2)
                if gov:
                    dist = district_qs.filter(parent_id=gov.id).first()

            cadaster_qs = Location.objects.none()
            cad = None
            if cad_name:
                cadaster_qs = Location.objects.filter(name_en=cad_name, type_id=3)
                if dist:
                    cad = cadaster_qs.filter(parent_id=dist.id).first()
            disability = Disability.objects.filter(name_en=values.get('disability')).first()

            if not nationality:
                invalid_fields.append("nationality ({0})".format(values.get('nationality')))
            if not gov:
                invalid_fields.append("governorate ({0})".format(values.get('governorate')))
            if not dist:
                if gov and district_qs.exists():
                    parent_ids = [pid for pid in district_qs.values_list('parent_id', flat=True) if pid]
                    parent_names = list(
                        Location.objects.filter(id__in=parent_ids).values_list('name_en', flat=True)
                    )
                    if parent_names:
                        parent_names = sorted(set(parent_names))
                        invalid_fields.append(
                            "district ({0}) belongs to governorate ({1}), not governorate ({2})".format(
                                values.get('district'),
                                ', '.join(parent_names),
                                values.get('governorate')
                            )
                        )
                    else:
                        invalid_fields.append(
                            "district ({0}) does not belong to governorate ({1})".format(
                                values.get('district'), values.get('governorate')
                            )
                        )
                else:
                    invalid_fields.append("district ({0})".format(values.get('district')))
            if not cad:
                if cadaster_qs.exists():
                    if dist:
                        invalid_fields.append(
                            "cadaster ({0}) does not belong to district ({1})".format(
                                values.get('cadaster'), values.get('district')
                            )
                        )
                    else:
                        parent_ids = [pid for pid in cadaster_qs.values_list('parent_id', flat=True) if pid]
                        parent_names = list(
                            Location.objects.filter(id__in=parent_ids).values_list('name_en', flat=True)
                        )
                        if parent_names:
                            parent_names = sorted(set(parent_names))
                            invalid_fields.append(
                                "cadaster ({0}) belongs to district ({1}), but district is invalid".format(
                                    values.get('cadaster'), ', '.join(parent_names)
                                )
                            )
                        else:
                            invalid_fields.append("cadaster ({0})".format(values.get('cadaster')))
                else:
                    invalid_fields.append("cadaster ({0})".format(values.get('cadaster')))
            if not disability:
                invalid_fields.append("disability ({0})".format(values.get('disability')))

            dob = None
            y = m = d = None
            try:
                y = int(values.get('birthday_year'))
                m = int(values.get('birthday_month'))
                d = int(values.get('birthday_day'))
                dob = datetime.date(y, m, d)
                if dob >= datetime.date.today():
                    invalid_fields.append("DOB ({0}) cannot be today or in the future".format(dob))
            except Exception:
                invalid_fields.append("DOB (invalid date parts)")

            # ---- Optional lookups: not mandatory, but if provided must exist

            ed_status = None
            ed_status_val = (values.get('education_status') or '').strip()
            if ed_status_val:
                valid_education_statuses = [choice[0] for choice in YOUTH_EDUCATION_STATUS]
                if ed_status_val in valid_education_statuses:
                    ed_status = ed_status_val
                else:
                    invalid_fields.append("education_status ({0})".format(ed_status_val))


            dropout_date = None
            dropout_date_val = values.get('dropout_date')
            if dropout_date_val:
                if isinstance(dropout_date_val, datetime.datetime):
                    dropout_date = dropout_date_val.date()
                elif isinstance(dropout_date_val, datetime.date):
                    dropout_date = dropout_date_val
                else:
                    dropout_date_val = str(dropout_date_val).strip()
                    try:
                        dropout_date = datetime.datetime.strptime(dropout_date_val, '%Y-%m-%d').date()
                        if dropout_date_val != dropout_date.strftime('%Y-%m-%d'):
                            invalid_fields.append("dropout_date ({0}) must be in yyyy-mm-dd format".format(dropout_date_val))
                    except Exception:
                        invalid_fields.append("dropout_date ({0}) must be in yyyy-mm-dd format".format(dropout_date_val))

            if (
                ed_status == 'Currently registered in Formal Education school but not attending'
                and not dropout_date
            ):
                invalid_fields.append("dropout_date is required when education_status is ({0})".format(ed_status))

            father_ed = None
            father_ed_val = (values.get('father_educational_level') or '').strip()
            if father_ed_val:
                father_ed = EducationalLevel.objects.filter(name_en=father_ed_val).first()
                if not father_ed:
                    invalid_fields.append("father_educational_level ({0})".format(father_ed_val))

            mother_ed = None
            mother_ed_val = (values.get('mother_educational_level') or '').strip()
            if mother_ed_val:
                mother_ed = EducationalLevel.objects.filter(name_en=mother_ed_val).first()
                if not mother_ed:
                    invalid_fields.append("mother_educational_level ({0})".format(mother_ed_val))

            caregiver_nat = None
            caregiver_nat_val = (values.get('main_caregiver_nationality') or '').strip()
            if caregiver_nat_val:
                caregiver_nat = Nationality.objects.filter(name_en=caregiver_nat_val).first()
                if not caregiver_nat:
                    invalid_fields.append("main_caregiver_nationality ({0})".format(caregiver_nat_val))

            id_type = None
            id_type_val = (values.get('id_type') or '').strip()
            if id_type_val:
                id_type = IDType.objects.filter(name=id_type_val).first()
                if not id_type:
                    invalid_fields.append("id_type ({0})".format(id_type_val))

            # ---- If any invalids, log and skip row
            if invalid_fields:
                not_imported.append(
                    self._build_error_entry(
                        values,
                        row_number,
                        "Invalid: " + "; ".join(invalid_fields)
                    )
                )
                continue

            values['gender'] = gender_norm

            validated_rows.append({
                'row_number': row_number,
                'values': values,
                'references': {
                    'nationality': nationality,
                    'gov': gov,
                    'dist': dist,
                    'cad': cad,
                    'disability': disability,
                    'ed_status': ed_status,
                    'dropout_date': dropout_date,
                    'father_ed': father_ed,
                    'mother_ed': mother_ed,
                    'caregiver_nat': caregiver_nat,
                    'id_type': id_type,
                },
                'dob': dob,
                'gender': gender_norm,
                'nationality_en': getattr(nationality, 'name_en', None) or getattr(nationality, 'name', ''),
            })

        # ---- Generate UNICEF IDs in bulk for validated rows
        payload = {
            "individuals": [
                {
                    "id": idx,
                    "first_name": row['values'].get('first_name') or '',
                    "father_name": row['values'].get('father_name') or '',
                    "last_name": row['values'].get('last_name') or '',
                    "mother_name": row['values'].get('mother_fullname') or '',
                    "date_of_birth": row['dob'].strftime('%Y-%m-%d') if row['dob'] else '',
                    "nationality": row['nationality_en'] or '',
                    "gender": row['gender'] or '',
                }
                for idx, row in enumerate(validated_rows)
            ]
        }

        bulk_ids = {}
        if payload["individuals"]:
            print("generate_bulk_unique_id sent:", payload)
            try:
                bulk_ids = generate_bulk_unique_id(payload) or {}
                print("generate_bulk_unique_id received:", bulk_ids)
            except Exception as exc:
                print("generate_bulk_unique_id received error:", exc)
                bulk_ids = {}

        # ---- Persist valid rows
        for idx, row in enumerate(validated_rows):
            values = row['values']
            references = row['references']
            row_number = row['row_number']
            prospective_unicef_id = bulk_ids.get(idx)


            if self._is_unicef_id_generation_error(prospective_unicef_id):
                not_imported.append(
                    self._build_error_entry(
                        values,
                        row_number,
                        self._unicef_id_generation_error_message(prospective_unicef_id)
                    )
                )
                continue

            if Registration.objects.filter(adolescent__unicef_id=prospective_unicef_id, deleted=False).exists():
                not_imported.append(
                    self._build_error_entry(
                        values,
                        row_number,
                        "Invalid: duplicate unicef_id ({0})".format(prospective_unicef_id)
                    )
                )
                continue

            try:
                adolescent = Adolescent.objects.create(
                    first_name=values.get('first_name'),
                    father_name=values.get('father_name'),
                    last_name=values.get('last_name'),
                    birthday_year=values.get('birthday_year'),
                    birthday_month=values.get('birthday_month'),
                    birthday_day=values.get('birthday_day'),
                    gender=values.get('gender'),
                    mother_fullname=values.get('mother_fullname'),
                    nationality=references['nationality'],
                    nationality_other=values.get('nationality_other'),
                    governorate=references['gov'],
                    district=references['dist'],
                    cadaster=references['cad'],
                    address=values.get('address'),
                    disability=references['disability'],
                    father_educational_level=references['father_ed'],
                    mother_educational_level=references['mother_ed'],
                    first_phone_number=values.get('first_phone_number'),
                    second_phone_number=values.get('second_phone_number'),
                    main_caregiver=values.get('main_caregiver'),
                    main_caregiver_other=values.get('main_caregiver_other'),
                    caregiver_first_name=values.get('caregiver_first_name'),
                    caregiver_middle_name=values.get('caregiver_middle_name'),
                    caregiver_last_name=values.get('caregiver_last_name'),
                    main_caregiver_nationality=references['caregiver_nat'],
                    main_caregiver_nationality_other=values.get('main_caregiver_nationality_other'),
                    id_type=references['id_type'],
                    case_number=values.get('case_number'),
                    parent_individual_case_number=values.get('parent_individual_case_number'),
                    individual_case_number=values.get('individual_case_number'),
                    recorded_number=values.get('recorded_number'),
                    unrwa_number=values.get('unrwa_number'),
                    parent_syrian_national_number=values.get('parent_syrian_national_number'),
                    syrian_national_number=values.get('syrian_national_number'),
                    parent_sop_national_number=values.get('parent_sop_national_number'),
                    sop_national_number=values.get('sop_national_number'),
                    parent_national_number=values.get('parent_national_number'),
                    national_number=values.get('national_number'),
                    parent_other_number=values.get('parent_other_number'),
                    other_number=values.get('other_number'),
                )

                adolescent.unicef_id = prospective_unicef_id
                adolescent.save()

                Registration.objects.create(
                    adolescent=adolescent,
                    education_status=references['ed_status'],
                    dropout_date=references['dropout_date'],
                    owner=request.user,
                    partner_id=getattr(request.user, 'partner_id', None),
                    center_id=getattr(request.user, 'center_id', None),
                )

                imported += 1

            except Exception as ex:
                not_imported.append(
                    self._build_error_entry(
                        values,
                        row_number,
                        str(ex)
                    )
                )

        # ---- Write failed rows CSV
        if not_imported:
            csv_buffer = io.StringIO()
            writer = csv.DictWriter(csv_buffer, fieldnames=not_imported[0].keys())
            writer.writeheader()
            writer.writerows(not_imported)
            upload.failed_file.save(
                'failed_{0}.csv'.format(upload.pk),
                ContentFile(csv_buffer.getvalue().encode('utf-8'))
            )

        return imported, not_imported


class AdolescentUploadFailedView(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        upload = get_object_or_404(AdolescentUpload, pk=pk, uploaded_by=request.user)
        if not upload.failed_file:
            return HttpResponse(status=404)

        file_content = upload.failed_file.read()
        bom = b'\xef\xbb\xbf'

        response = HttpResponse(bom + file_content, content_type='text/csv; charset=utf-8')
        filename = upload.failed_file.name.split('/')[-1]
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename

        return response

