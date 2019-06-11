import json

from django.core.serializers.json import DjangoJSONEncoder
from openpyxl import load_workbook

from utils import post_data
from student_registration.taskapp.celery import app


@app.task
def push_bln_data(file_name, base_url, token, protocol='HTTPS'):

    wb = load_workbook(filename='test.XLSX', read_only=True)
    ws = wb['bln']

    for row in ws.rows:
        if row[0].value == 'skip':
            continue

        try:
            data = {}
            data['round'] = row[1].value if row[1].value else 1
            data['new_registry'] = row[2].value if row[2].value else 'yes'
            data['student_outreached'] = row[3].value if row[3].value else 'no'
            data['have_barcode'] = row[4].value if row[4].value else 'no'
            data['partner'] = 10
            data['governorate'] = row[6].value
            data['district'] = row[7].value
            # data['location'] = row[8].value
            data['language'] = row[8].value if row[8].value else 'english_arabic'
            data['student_first_name'] = row[9].value if row[9].value else 'None'
            data['student_father_name'] = row[10].value if row[10].value else 'None'
            data['student_mother_fullname'] = row[11].value if row[11].value else 'None'
            data['student_last_name'] = row[12].value if row[12].value else 'None'
            data['student_sex'] = row[13].value if row[13].value else 'Male'
            data['student_nationality'] = row[14].value if row[14].value else 1
            data['student_birthday_year'] = row[15].value if row[15].value else '2009'
            data['student_birthday_month'] = row[16].value if row[16].value else '1'
            data['student_birthday_day'] = row[17].value if row[17].value else '1'
            data['student_p_code'] = row[18].value if row[18].value else 'None'
            data['disability'] = row[19].value if row[19].value else 1
            data['student_id_number'] =  row[20].value if row[20].value else 'None'

            if row[21].value:
                internal_number = row[21].value
            else:
                internal_number = 'None'

            data['internal_number'] = internal_number

            data['comments'] = row[22].value if row[22].value else 'None'
            data['hh_educational_level'] = row[23].value
            data['student_family_status'] = row[24].value if row[24].value else 'single'
            data['student_have_children'] = row[25].value if row[25].value else 0
            data['have_labour'] = [row[26].value] if row[26].value else []
            # data['labours'] = row[27].value
            data['labour_hours'] = row[28].value
            data['pre_test_language'] = row[29].value
            data['pre_test_math'] = row[30].value
            data['pre_test_arabic'] = row[31].value
            data['post_test_language'] = row[32].value
            data['post_test_math'] = row[33].value
            data['post_test_arabic'] = row[34].value

            if row[35].value:
                data['unsuccessful_posttest_reason'] = row[35].value
            data['participation'] = row[36].value

            # data['barriers'] = [''.join(row[37].value.split())] if (row[37].value or row[37].value == ' ') else []
            data['learning_result'] = row[38].value

            result = post_data(protocol=protocol, url=base_url, apifunc='/api/clm-bln/', token=token, data=data)
            result = json.loads(result)
            # print(sresult)

        except Exception as ex:
            print("---------------1")
            print("error: ", ex.message)
            print(json.dumps(data, cls=DjangoJSONEncoder))
            print("---------------")
            continue


@app.task
def push_cbece_data(file_name, base_url, token, protocol='HTTPS'):

    wb = load_workbook(filename='test.XLSX', read_only=True)
    ws = wb['cbece']

    for row in ws.rows:
        if row[0].value == 'skip':
            continue

        try:
            data = {}
            data['partner'] = 10
            data['round'] = row[1].value if row[1].value else 1
            data['new_registry'] = row[2].value if row[2].value else 'yes'
            data['student_outreached'] = row[3].value if row[3].value else 'no'
            data['have_barcode'] = row[4].value if row[4].value else 'no'
            cycle = row[5].value if row[5].value else 1
            data['cycle'] = cycle
            data['id_site'] = row[6].value if row[6].value else 1
            data['school'] = row[7].value
            data['governorate'] = row[8].value
            data['district'] = row[9].value
            data['location'] = row[10].value
            data['language'] = row[11].value if row[11].value else 'english_arabic'
            data['referral'] = [row[12].value] if row[12].value else []
            data['student_first_name'] = row[13].value if row[13].value else 'None'
            data['student_father_name'] = row[14].value if row[14].value else 'None'
            data['student_mother_fullname'] = row[15].value if row[15].value else 'None'
            data['student_last_name'] = row[16].value if row[16].value else 'None'
            data['student_sex'] = row[17].value if row[17].value else 'Male'
            data['student_nationality'] = row[18].value if row[18].value else 1
            data['student_birthday_year'] = row[19].value if row[19].value else '2009'
            data['student_birthday_month'] = row[20].value if row[20].value else '1'
            data['student_birthday_day'] = row[21].value if row[21].value else '1'
            data['student_p_code'] = row[22].value if row[22].value else 'None'
            data['disability'] = row[23].value if row[23].value else 1
            data['student_id_number'] = row[24].value if row[24].value else 'None'
            data['internal_number'] = row[25].value if row[25].value else 'None'
            data['comments'] = row[26].value if row[26].value else 'None'
            data['child_muac'] = row[27].value if row[27].value else 2
            data['hh_educational_level'] = row[28].value
            data['have_labour'] = [row[29].value] if row[29].value else []
            data['labours'] = row[30].value
            data['labour_hours'] = row[31].value

            if row[44].value:
                data['unsuccessful_posttest_reason'] = row[44].value
            data['participation'] = row[45].value
            data['barriers'] = [''.join(row[46].value.split())] if (row[46].value or row[46].value == ' ') else []
            data['learning_result'] = row[47].value

            result = post_data(protocol=protocol, url=base_url, apifunc='/api/clm-cbece/', token=token, data=data)
            result = json.loads(result)
            # print(sresult)

            # pre test
            try:
                assessment = {}
                assessment['status'] = 'pre_test'
                assessment['enrollment_id'] = result['original_id']
                assessment['enrollment_model'] = 'CBECE'
                assessment['CBECE_ASSESSMENT/LanguageArtDomain' + str(cycle)] = row[32].value if row[32].value else '0'
                assessment['CBECE_ASSESSMENT/CognitiveDomian' + str(cycle)] = row[33].value if row[33].value else '0'
                assessment['CBECE_ASSESSMENT/ScienceDomain' + str(cycle)] = row[34].value if row[34].value else '0'
                assessment['CBECE_ASSESSMENT/SocialEmotionalDomain' + str(cycle)] = row[35].value if row[35].value else '0'
                assessment['CBECE_ASSESSMENT/PsychomotorDomain'+ str(cycle)] = row[36].value if row[36].value else '0'
                assessment['CBECE_ASSESSMENT/ArtisticDomain' + str(cycle)] = row[37].value if row[37].value else '0'

                post_data(protocol=protocol, url=base_url, apifunc='/clm/assessment-submission/', token=token,
                          data=assessment)
            except Exception as ex:
                print("---------------1")
                print("error: ", ex.message)
                print(json.dumps(assessment, cls=DjangoJSONEncoder))
                print("---------------")
                pass

            # post test
            try:
                assessment = {}
                assessment['status'] = 'post_test'
                assessment['enrollment_id'] = result['original_id']
                assessment['enrollment_model'] = 'CBECE'
                assessment['CBECE_ASSESSMENT/LanguageArtDomain' + str(cycle)] = row[38].value if row[38].value else '0'
                assessment['CBECE_ASSESSMENT/CognitiveDomian' + str(cycle)] = row[39].value if row[39].value else '0'
                assessment['CBECE_ASSESSMENT/ScienceDomain' + str(cycle)] = row[40].value if row[40].value else '0'
                assessment['CBECE_ASSESSMENT/SocialEmotionalDomain' + str(cycle)] = row[41].value if row[41].value else '0'
                assessment['CBECE_ASSESSMENT/PsychomotorDomain' + str(cycle)] = row[42].value if row[42].value else '0'
                assessment['CBECE_ASSESSMENT/ArtisticDomain' + str(cycle)] = row[43].value if row[43].value else '0'

                post_data(protocol=protocol, url=base_url, apifunc='/clm/assessment-submission/', token=token,
                          data=assessment)
            except Exception as ex:
                print("---------------2")
                print("error: ", ex.message)
                print(json.dumps(assessment, cls=DjangoJSONEncoder))
                print("---------------")
                pass

        except Exception as ex:
            print("---------------1")
            print("error: ", ex.message)
            print(json.dumps(data, cls=DjangoJSONEncoder))
            print("---------------")
            continue


@app.task
def push_rs_data(file_name, base_url, token, protocol='HTTPS'):

    # from .models import RS
    # objects = RS.objects.all()
    # objects.delete()

    wb = load_workbook(filename='test.XLSX', read_only=True)
    ws = wb['RS']

    for row in ws.rows:
        if row[0].value == 'skip':
            continue

        try:
            data = {}

            data['student_outreached'] = 'no'
            data['have_barcode'] = 'no'
            data['partner'] = 10
            # data['partner'] = row[0].value
            data['round'] = row[1].value if row[1].value else 4
            data['new_registry'] = 'yes'
            # data['type'] = row[3].value if row[3].value else 'homework_support'
            data['type'] = 'homework_support'
            # data['site'] = row[4].value if row[4].value else 'in_school'
            data['site'] = 'in_school'

            data['school'] = row[5].value
            data['governorate'] = row[6].value
            data['district'] = row[7].value
            data['location'] = row[8].value
            data['language'] = row[9].value if row[9].value else 'english_arabic'
            data['student_first_name'] = row[10].value if row[10].value else 'None'
            data['student_father_name'] = row[11].value if row[11].value else 'None'
            data['student_mother_fullname'] = row[12].value if row[12].value else 'None'
            data['student_last_name'] = row[13].value if row[13].value else 'None'
            data['student_sex'] = row[14].value if row[14].value else 'Male'
            data['student_nationality'] = row[15].value if row[15].value else 1
            data['student_birthday_year'] = row[16].value if row[16].value else '2009'
            data['student_birthday_month'] = row[17].value if row[17].value else '1'
            data['student_birthday_day'] = row[18].value if row[18].value else '1'
            data['student_p_code'] = row[19].value if row[19].value else 'None'
            data['disability'] = row[20].value if row[20].value else 1
            data['student_id_number'] = row[21].value if row[21].value else 'None'
            data['internal_number'] = row[22].value if row[22].value else 'None'
            data['comments'] = row[23].value if row[23].value else 'None'
            data['hh_educational_level'] = row[27].value
            data['student_family_status'] = row[28].value if row[28].value else 'single'
            data['student_have_children'] = row[29].value if row[29].value else 0
            data['have_labour'] = [row[30].value] if row[30].value else []
            data['labours'] = row[31].value
            data['labour_hours'] = row[32].value
            data['registered_in_school'] = row[33].value
            data['shift'] = row[34].value if row[34].value else 'first'
            data['grade'] = row[35].value
            data['section'] = row[36].value
            data['referral'] = [row[37].value] if row[37].value else []
            data['pre_test_arabic'] = row[38].value
            data['pre_test_language'] = row[39].value
            data['pre_test_math'] = row[40].value
            data['pre_test_science'] = row[41].value
            if row[92].value:
                data['unsuccessful_posttest_reason'] = row[92].value
            data['participation'] = row[93].value
            data['barriers'] = [''.join(row[94].value.split())] if (row[94].value or row[94].value == ' ') else []
            data['learning_result'] = row[95].value
            data['post_test_arabic'] = row[65].value
            data['post_test_language'] = row[66].value
            data['post_test_math'] = row[67].value
            data['post_test_science'] = row[68].value

            result = post_data(protocol=protocol, url=base_url, apifunc='/api/clm-rs/', token=token, data=data)
            result = json.loads(result)
            # print(result)

            # pre reading
            try:
                assessment = {}
                assessment['status'] = 'pre_reading'
                assessment['enrollment_id'] = result['original_id']
                assessment['enrollment_model'] = 'RS'
                assessment['RS_ASSESSMENT/FL1'] = row[42].value if row[42].value else '2'
                post_data(protocol=protocol, url=base_url, apifunc='/clm/assessment-submission/', token=token,
                          data=assessment)
            except Exception as ex:
                print("---------------1")
                print("error: ", ex.message)
                print (assessment)
                print(json.dumps(assessment, cls=DjangoJSONEncoder))
                print("---------------1")
                pass



            # post reading
            try:
                assessment = {}
                assessment['status'] = 'post_reading'
                assessment['enrollment_id'] = result['original_id']
                assessment['enrollment_model'] = 'RS'
                assessment['RS_ASSESSMENT/FL1'] = row[69].value if row[69].value else '2'

                post_data(protocol=protocol, url=base_url, apifunc='/clm/assessment-submission/', token=token,
                          data=assessment)
            except Exception as ex:
                print("---------------2")
                print("error: ", ex.message)
                print(json.dumps(assessment, cls=DjangoJSONEncoder))
                print("---------------")
                pass



            # pre test
            try:
                assessment = {}
                assessment['status'] = 'pre_test'
                assessment['enrollment_id'] = result['original_id']
                assessment['enrollment_model'] = 'RS'
                assessment['RS_ASSESSMENT/FL1'] = row[43].value if row[43].value else '0'
                assessment['RS_ASSESSMENT/FL2'] = row[44].value if row[44].value else '0'
                assessment['RS_ASSESSMENT/FL3'] = row[45].value if row[45].value else '0'
                assessment['RS_ASSESSMENT/FL4'] = row[46].value if row[46].value else '0'

                post_data(protocol=protocol, url=base_url, apifunc='/clm/assessment-submission/', token=token,
                          data=assessment)
            except Exception as ex:
                print("---------------3")
                print("error: ", ex.message)
                print(json.dumps(assessment, cls=DjangoJSONEncoder))
                print("---------------")
                pass

            # post test
            try:
                assessment = {}
                assessment['status'] = 'post_test'
                assessment['enrollment_id'] = result['original_id']
                assessment['enrollment_model'] = 'RS'
                assessment['RS_ASSESSMENT/FL1'] = row[70].value if row[70].value else '0'
                assessment['RS_ASSESSMENT/FL2'] = row[71].value if row[71].value else '0'
                assessment['RS_ASSESSMENT/FL3'] = row[72].value if row[72].value else '0'
                assessment['RS_ASSESSMENT/FL4'] = row[73].value if row[73].value else '0'

                post_data(protocol=protocol, url=base_url, apifunc='/clm/assessment-submission/', token=token,
                          data=assessment)
            except Exception as ex:
                print("---------------4")
                print("error: ", ex.message)
                print(json.dumps(assessment, cls=DjangoJSONEncoder))
                print("---------------")
                pass

            # pre_motivation
            try:
                assessment = {}
                assessment['status'] = 'pre_motivation'
                assessment['enrollment_id'] = result['original_id']
                assessment['enrollment_model'] = 'RS'
                assessment['RS_ASSESSMENT/FL5'] = row[47].value if row[47].value else '1'
                assessment['RS_ASSESSMENT/FL6'] = row[48].value if row[48].value else '1'
                assessment['RS_ASSESSMENT/FL7'] = row[49].value if row[49].value else '1'
                assessment['RS_ASSESSMENT/FL8'] = row[50].value if row[50].value else '1'

                post_data(protocol=protocol, url=base_url, apifunc='/clm/assessment-submission/', token=token,
                          data=assessment)
            except Exception as ex:
                print("---------------5")
                print("error: ", ex.message)
                print(json.dumps(assessment, cls=DjangoJSONEncoder))
                print("---------------")
                pass
            # post_motivation
            try:
                assessment = {}
                assessment['status'] = 'post_motivation'
                assessment['enrollment_id'] = result['original_id']
                assessment['enrollment_model'] = 'RS'
                assessment['RS_ASSESSMENT/FL5'] = row[74].value if row[74].value else '1'
                assessment['RS_ASSESSMENT/FL6'] = row[75].value if row[75].value else '1'
                assessment['RS_ASSESSMENT/FL7'] = row[76].value if row[76].value else '1'
                assessment['RS_ASSESSMENT/FL8'] = row[77].value if row[77].value else '1'

                post_data(protocol=protocol, url=base_url, apifunc='/clm/assessment-submission/', token=token,
                          data=assessment)
            except Exception as ex:
                print("---------------6")
                print("error: ", ex.message)
                print(json.dumps(assessment, cls=DjangoJSONEncoder))
                print("---------------")
                pass
            # pre_self_assessment
            try:
                assessment = {}
                assessment['status'] = 'pre_self_assessment'
                assessment['enrollment_id'] = result['original_id']
                assessment['enrollment_model'] = 'RS'
                assessment['SELF_ASSESSMENT/assessment_1'] = row[51].value if row[51].value else '0'
                assessment['SELF_ASSESSMENT/assessment_2'] = row[52].value if row[52].value else '0'
                assessment['SELF_ASSESSMENT/assessment_3'] = row[53].value if row[53].value else '0'
                assessment['SELF_ASSESSMENT/assessment_4'] = row[54].value if row[54].value else '0'
                assessment['SELF_ASSESSMENT/assessment_5'] = row[55].value if row[55].value else '0'
                assessment['SELF_ASSESSMENT/assessment_6'] = row[56].value if row[56].value else '0'
                assessment['SELF_ASSESSMENT/assessment_7'] = row[57].value if row[57].value else '0'
                assessment['SELF_ASSESSMENT/assessment_8'] = row[58].value if row[58].value else '0'
                assessment['SELF_ASSESSMENT/assessment_9'] = row[59].value if row[59].value else '0'
                assessment['SELF_ASSESSMENT/assessment_10'] = row[60].value if row[60].value else '0'
                assessment['SELF_ASSESSMENT/assessment_11'] = row[61].value if row[61].value else '0'
                assessment['SELF_ASSESSMENT/assessment_12'] = row[62].value if row[62].value else '0'
                assessment['SELF_ASSESSMENT/assessment_13'] = row[63].value if row[63].value else '0'
                assessment['SELF_ASSESSMENT/assessment_14'] = row[64].value if row[64].value else '0'
                post_data(protocol=protocol, url=base_url, apifunc='/clm/assessment-submission/', token=token,
                          data=assessment)
            except Exception as ex:
                print("---------------7")
                print("error: ", ex.message)
                print(json.dumps(assessment, cls=DjangoJSONEncoder))
                print("---------------")
                pass

            # post_self_assessment
            try:
                assessment = {}
                assessment['status'] = 'post_self_assessment'
                assessment['enrollment_id'] = result['original_id']
                assessment['enrollment_model'] = 'RS'
                assessment['SELF_ASSESSMENT/assessment_1'] = row[78].value if row[78].value else '0'
                assessment['SELF_ASSESSMENT/assessment_2'] = row[79].value if row[79].value else '0'
                assessment['SELF_ASSESSMENT/assessment_3'] = row[80].value if row[80].value else '0'
                assessment['SELF_ASSESSMENT/assessment_4'] = row[81].value if row[81].value else '0'
                assessment['SELF_ASSESSMENT/assessment_5'] = row[82].value if row[82].value else '0'
                assessment['SELF_ASSESSMENT/assessment_6'] = row[83].value if row[83].value else '0'
                assessment['SELF_ASSESSMENT/assessment_7'] = row[84].value if row[84].value else '0'
                assessment['SELF_ASSESSMENT/assessment_8'] = row[85].value if row[85].value else '0'
                assessment['SELF_ASSESSMENT/assessment_9'] = row[86].value if row[86].value else '0'
                assessment['SELF_ASSESSMENT/assessment_10'] = row[87].value if row[87].value else '0'
                assessment['SELF_ASSESSMENT/assessment_11'] = row[88].value if row[88].value else '0'
                assessment['SELF_ASSESSMENT/assessment_12'] = row[89].value if row[89].value else '0'
                assessment['SELF_ASSESSMENT/assessment_13'] = row[90].value if row[90].value else '0'
                assessment['SELF_ASSESSMENT/assessment_14'] = row[91].value if row[91].value else '0'
                post_data(protocol=protocol, url=base_url, apifunc='/clm/assessment-submission/', token=token,
                          data=assessment)
            except Exception as ex:
                print("---------------8")
                print("error: ", ex.message)
                print(json.dumps(assessment, cls=DjangoJSONEncoder))
                print("---------------")
                pass
        except Exception as ex:
            print("---------------9")
            print("error: ", ex.message)
            print(json.dumps(data, cls=DjangoJSONEncoder))
            print("---------------")
            continue

