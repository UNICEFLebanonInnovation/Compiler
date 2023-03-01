# -- coding: utf-8 --
import io
import xlwt
import csv
from datetime import date
from django.http import HttpResponse, FileResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color

import copy

import logging
logger = logging.getLogger(__name__)

from datetime import datetime

from .models import Bridging
from student_registration.students.models import Person
from student_registration.outreach.models import OutreachChild


def is_allowed_create(programme):
    from student_registration.schools.models import CLMRound
    try:
        current = date.today()
        current_round = CLMRound.objects.all()

        if programme == 'BLN':
            current_round = current_round.get(current_round_bln=True)
            if current_round.start_date_bln < current < current_round.end_date_bln:
                return True
            return False

        if programme == 'ABLN':
            current_round = current_round.get(current_round_abln=True)
            if current_round.start_date_abln < current < current_round.end_date_abln:
                return True
            return False

        if programme == 'CBECE':
            current_round = current_round.get(current_round_cbece=True)
            if current_round.start_date_cbece < current < current_round.end_date_cbece:
                return True
            return False

        if programme == 'Inclusion':
            current_round = current_round.get(current_round_inclusion=True)
            if current_round.start_date_inclusion < current < current_round.end_date_inclusion:
                return True
            return False

        if programme == 'RS':
            current_round = current_round.get(current_round_rs=True)
            if current_round.start_date_rs < current < current_round.end_date_rs:
                return True
            return False

        if programme == 'Outreach':
            current_round = current_round.get(current_round_outreach=True)
            if current_round.start_date_outreach < current < current_round.end_date_outreach:
                return True
            return False

        if programme == 'GeneralQuestionnaire':
            return True

        if programme == 'Bridging':
            current_round = current_round.get(current_round_bridging=True)
            if current_round.start_date_bridging < current < current_round.end_date_bridging:
                return True
            return False

        if programme == 'MSCC':
            current_round = current_round.get(current_round_mscc=True)
            if current_round.start_date_mscc < current < current_round.end_date_mscc:
                return True
            return False


    except Exception as ex:
        print(ex.message)
        return False


def is_allowed_edit(programme):
    from student_registration.schools.models import CLMRound

    try:
        current = date.today()
        current_round = CLMRound.objects.all()

        if programme == 'BLN':
            current_round = current_round.get(current_round_bln=True)
            if current_round.start_date_bln_edit < current < current_round.end_date_bln_edit:
                return True
            return False

        if programme == 'ABLN':
            current_round = current_round.get(current_round_abln=True)
            if current_round.start_date_abln_edit < current < current_round.end_date_abln_edit:
                return True
            return False

        if programme == 'CBECE':
            current_round = current_round.get(current_round_cbece=True)
            if current_round.start_date_cbece_edit < current < current_round.end_date_cbece_edit:
                return True
            return False

        if programme == 'Inclusion':
            current_round = current_round.get(current_round_inclusion=True)
            if current_round.start_date_inclusion_edit < current < current_round.end_date_inclusion_edit:
                return True
            return False

        if programme == 'RS':
            current_round = current_round.get(current_round_rs=True)
            if current_round.start_date_rs_edit < current < current_round.end_date_rs_edit:
                return True
            return False

        if programme == 'Outreach':
            current_round = current_round.get(current_round_outreach=True)
            if current_round.start_date_outreach_edit < current < current_round.end_date_outreach_edit:
                return True
            return False

        if programme == 'GeneralQuestionnaire':
            return True

        if programme == 'Bridging':
            current_round = current_round.get(current_round_bridging=True)
            if current_round.start_date_bridging_edit < current < current_round.end_date_bridging_edit:
                return True
            return False

        if programme == 'MSCC':
            current_round = current_round.get(current_round_bridging=True)
            if current_round.start_date_mscc_edit < current < current_round.end_date_mscc_edit:
                return True
            return False

    except Exception as ex:
        print(ex.message)
        return False


def bln_build_xls_extraction(queryset_students, queryset_fc):
    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        'enrollment_id',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'first attendance date',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'RIMS  case number',
        'Source of Identification',
        'Source of Identification Specify',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'main caregiver nationality Other ',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'test_done',
        'round_complete',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Did the child use Akelius program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session number',
        'pss session modality',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'pre test attended arabic',
        'pre test modality arabic',
        'pre test arabic',
        'pre test attended english',
        'pre test modality english',
        'pre test english',
        'pre test attended psychomotor',
        'pre test modality psychomotor',
        'pre test psychomotor',
        'pre test attended artistic',
        'pre test modality artistic',
        'pre test artistic',
        'pre test attended math',
        'pre test modality math',
        'pre test math',
        'pre test attended social',
        'pre test modality social',
        'pre test social emotional',
        'post test attended arabic',
        'post test modality arabic',
        'post test arabic',
        'post test attended english',
        'post test modality english',
        'post test english',
        'post test attended psychomotor',
        'post test modality psychomotor',
        'post test psychomotor',
        'post test attended artistic',
        'post test modality artistic',
        'post test artistic',
        'post test attended math',
        'post test modality math',
        'post test math',
        'post test attended social',
        'post test modality social',
        'post test social emotional',
        'owner',
        'modified_by',
    ]

    for col_num in range(len(columns)):
        ws.write(row_num_student, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    qs = queryset_students.extra(select={

        'pre_test_attended_arabic': "pre_test->>'BLN_ASSESSMENT/attended_arabic'",
        'pre_test_modality_arabic': "pre_test->>'BLN_ASSESSMENT/modality_arabic'",
        'pre_test_arabic': "pre_test->>'BLN_ASSESSMENT/arabic'",

        'pre_test_attended_english': "pre_test->>'BLN_ASSESSMENT/attended_english'",
        'pre_test_modality_english': "pre_test->>'BLN_ASSESSMENT/modality_english'",
        'pre_test_english': "pre_test->>'BLN_ASSESSMENT/english'",

        'pre_test_attended_psychomotor': "pre_test->>'BLN_ASSESSMENT/attended_psychomotor'",
        'pre_test_modality_psychomotor': "pre_test->>'BLN_ASSESSMENT/modality_psychomotor'",
        'pre_test_psychomotor': "pre_test->>'BLN_ASSESSMENT/psychomotor'",

        'pre_test_attended_artistic': "pre_test->>'BLN_ASSESSMENT/attended_artistic'",
        'pre_test_modality_artistic': "pre_test->>'BLN_ASSESSMENT/modality_artistic'",
        'pre_test_artistic': "pre_test->>'BLN_ASSESSMENT/artistic'",

        'pre_test_attended_math': "pre_test->>'BLN_ASSESSMENT/attended_math'",
        'pre_test_modality_math': "pre_test->>'BLN_ASSESSMENT/modality_math'",
        'pre_test_math': "pre_test->>'BLN_ASSESSMENT/math'",

        'pre_test_attended_social': "pre_test->>'BLN_ASSESSMENT/attended_social'",
        'pre_test_modality_social': "pre_test->>'BLN_ASSESSMENT/modality_social'",
        'pre_test_social_emotional': "pre_test->>'BLN_ASSESSMENT/social_emotional'",

        'post_test_attended_arabic': "post_test->>'BLN_ASSESSMENT/attended_arabic'",
        'post_test_modality_arabic': "post_test->>'BLN_ASSESSMENT/modality_arabic'",
        'post_test_arabic': "post_test->>'BLN_ASSESSMENT/arabic'",

        'post_test_attended_english': "post_test->>'BLN_ASSESSMENT/attended_english'",
        'post_test_modality_english': "post_test->>'BLN_ASSESSMENT/modality_english'",
        'post_test_english': "post_test->>'BLN_ASSESSMENT/english'",

        'post_test_attended_psychomotor': "post_test->>'BLN_ASSESSMENT/attended_psychomotor'",
        'post_test_modality_psychomotor': "post_test->>'BLN_ASSESSMENT/modality_psychomotor'",
        'post_test_psychomotor': "post_test->>'BLN_ASSESSMENT/psychomotor'",

        'post_test_attended_artistic': "post_test->>'BLN_ASSESSMENT/attended_artistic'",
        'post_test_modality_artistic': "post_test->>'BLN_ASSESSMENT/modality_artistic'",
        'post_test_artistic': "post_test->>'BLN_ASSESSMENT/artistic'",

        'post_test_attended_math': "post_test->>'BLN_ASSESSMENT/attended_math'",
        'post_test_modality_math': "post_test->>'BLN_ASSESSMENT/modality_math'",
        'post_test_math': "post_test->>'BLN_ASSESSMENT/math'",

        'post_test_attended_social': "post_test->>'BLN_ASSESSMENT/attended_social'",
        'post_test_modality_social': "post_test->>'BLN_ASSESSMENT/modality_social'",
        'post_test_social_emotional': "post_test->>'BLN_ASSESSMENT/social_emotional'",
    })

    rows = qs.order_by('id').values_list(
        'id',
        'new_registry',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'rims_case_number',
        'source_of_identification',
        'source_of_identification_specify',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'main_caregiver_nationality_other',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'participation',
        'barriers_single',
        'barriers_other',
        'test_done',
        'round_complete',
        'basic_stationery',
        'pss_kit',
        'learning_result',
        'learning_result_other',
        'akelius_program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_number',
        'pss_session_modality',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_number',
        'covid_session_modality',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_number',
        'followup_session_modality',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'pre_test_attended_arabic',
        'pre_test_modality_arabic',
        'pre_test_arabic',
        'pre_test_attended_english',
        'pre_test_modality_english',
        'pre_test_english',
        'pre_test_attended_psychomotor',
        'pre_test_modality_psychomotor',
        'pre_test_psychomotor',
        'pre_test_attended_artistic',
        'pre_test_modality_artistic',
        'pre_test_artistic',
        'pre_test_attended_math',
        'pre_test_modality_math',
        'pre_test_math',
        'pre_test_attended_social',
        'pre_test_modality_social',
        'pre_test_social_emotional',
        'post_test_attended_arabic',
        'post_test_modality_arabic',
        'post_test_arabic',
        'post_test_attended_english',
        'post_test_modality_english',
        'post_test_english',
        'post_test_attended_psychomotor',
        'post_test_modality_psychomotor',
        'post_test_psychomotor',
        'post_test_attended_artistic',
        'post_test_modality_artistic',
        'post_test_artistic',
        'post_test_attended_math',
        'post_test_modality_math',
        'post_test_math',
        'post_test_attended_social',
        'post_test_modality_social',
        'post_test_social_emotional',
        'owner__username',
        'modified_by__username',
    )
    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            ws.write(row_num_student, col_num, row[col_num], font_style)
    # FC
    wsFC = wb_student.add_sheet('FC')
    row_num_fc = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    column_header_fc = [
        'enrollment id',
        'Partner',
        'CLM Round',
        'First time registered?',
        'Registration level',
        'Student first name',
        'Student father name',
        'Student last name',
        'unique number',
        'Gender',
        'Student Nationality',
        'Does the child participate in work?',
        'Does the child have any disability or special need?',
        'fc type',
        'facilitator name',
        'subject taught',
        'date of monitoring',
        'targeted competencies',
        'activities reported',
        'activities reported other',
        'share expectations',
        'share expectations no reason',
        'share expectations other reason',
        'materials needed available',
        'attend lesson',
        'child interact teacher',
        'child interact friends',
        'child clear responses',
        'child ask questions',
        'child acquire competency',
        'child show improvement',
        'child expected work independently',
        'work independently evaluation',
        'complete printed package',
        'sessions participated',
        'not participating reason',
        'E recharge card provided',
        'action to taken',
        'action to taken specify',
        'child needs pss',
        'child cant access resources',
        'homework after lesson',
        'parents supporting student',
        'completed tasks',
        'meet objectives',
        'meet objectives verified',
        'objectives verified specify',
        'additional notes'
    ]

    for col_num in range(len(column_header_fc)):
        wsFC.write(row_num_fc, col_num, column_header_fc[col_num], font_style)
    font_style = xlwt.XFStyle()

    rows_fc = queryset_fc.values_list(

        'enrollment_id',
        'enrollment__partner__name',
        'enrollment__round__name',
        'enrollment__new_registry',
        'enrollment__registration_level',
        'enrollment__student__first_name',
        'enrollment__student__father_name',
        'enrollment__student__last_name',
        'enrollment__student__number',
        'enrollment__student__sex',
        'enrollment__student__nationality__name',
        'enrollment__disability__name_en',
        'enrollment__have_labour_single_selection',
        'fc_type',
        'facilitator_name',
        'subject_taught',
        'date_of_monitoring',
        'targeted_competencies',
        'activities_reported',
        'activities_reported_other',
        'share_expectations',
        'share_expectations_no_reason',
        'share_expectations_other_reason',
        'materials_needed_available',
        'attend_lesson',
        'child_interact_teacher',
        'child_interact_friends',
        'child_clear_responses',
        'child_ask_questions',
        'child_acquire_competency',
        'child_show_improvement',
        'child_expected_work_independently',
        'work_independently_evaluation',
        'complete_printed_package',
        'sessions_participated',
        'not_participating_reason',
        'e_recharge_card_provided',
        'action_to_taken',
        'action_to_taken_specify',
        'child_needs_pss',
        'child_cant_access_resources',
        'homework_after_lesson',
        'parents_supporting_student',
        'completed_tasks',
        'meet_objectives',
        'meet_objectives_verified',
        'objectives_verified_specify',
        'additional_notes'

    )

    for row in rows_fc:
        row_num_fc += 1
        for col_num in range(len(row)):
            wsFC.write(row_num_fc, col_num, row[col_num], font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="BLN.xls"'

    return response


def abln_build_xls_extraction(queryset_students, queryset_fc):
    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        'enrollment_id',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'first attendance date',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'RIMS  case number',
        'Source of Identification',
        'Source of Identification Specify',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'main caregiver nationality Other ',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'test_done',
        'round_complete',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Did the child use Akelius program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session number',
        'pss session modality',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'pre test attended arabic',
        'pre test modality arabic',
        'pre test arabic',
        'pre test attended artistic',
        'pre test modality artistic',
        'pre test artistic',
        'pre test attended math',
        'pre test modality math',
        'pre test math',
        'pre test attended social',
        'pre test modality social',
        'pre test social emotional',
        'post test attended arabic',
        'post test modality arabic',
        'post test arabic',
        'post test attended artistic',
        'post test modality artistic',
        'post test artistic',
        'post test attended math',
        'post test modality math',
        'post test math',
        'post test attended social',
        'post test modality social',
        'post test social emotional',
        'owner',
        'modified_by',
    ]

    for col_num in range(len(columns)):
        ws.write(row_num_student, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    qs = queryset_students.extra(select={

        'pre_test_attended_arabic': "pre_test->>'ABLN_ASSESSMENT/attended_arabic'",
        'pre_test_modality_arabic': "pre_test->>'ABLN_ASSESSMENT/modality_arabic'",
        'pre_test_arabic': "pre_test->>'ABLN_ASSESSMENT/arabic'",

        'pre_test_attended_artistic': "pre_test->>'ABLN_ASSESSMENT/attended_artistic'",
        'pre_test_modality_artistic': "pre_test->>'ABLN_ASSESSMENT/modality_artistic'",
        'pre_test_artistic': "pre_test->>'ABLN_ASSESSMENT/artistic'",

        'pre_test_attended_math': "pre_test->>'ABLN_ASSESSMENT/attended_math'",
        'pre_test_modality_math': "pre_test->>'ABLN_ASSESSMENT/modality_math'",
        'pre_test_math': "pre_test->>'ABLN_ASSESSMENT/math'",

        'pre_test_attended_social': "pre_test->>'ABLN_ASSESSMENT/attended_social'",
        'pre_test_modality_social': "pre_test->>'ABLN_ASSESSMENT/modality_social'",
        'pre_test_social_emotional': "pre_test->>'ABLN_ASSESSMENT/social_emotional'",

        'post_test_attended_arabic': "post_test->>'ABLN_ASSESSMENT/attended_arabic'",
        'post_test_modality_arabic': "post_test->>'ABLN_ASSESSMENT/modality_arabic'",
        'post_test_arabic': "post_test->>'ABLN_ASSESSMENT/arabic'",

        'post_test_attended_artistic': "post_test->>'ABLN_ASSESSMENT/attended_artistic'",
        'post_test_modality_artistic': "post_test->>'ABLN_ASSESSMENT/modality_artistic'",
        'post_test_artistic': "post_test->>'ABLN_ASSESSMENT/artistic'",

        'post_test_attended_math': "post_test->>'ABLN_ASSESSMENT/attended_math'",
        'post_test_modality_math': "post_test->>'ABLN_ASSESSMENT/modality_math'",
        'post_test_math': "post_test->>'ABLN_ASSESSMENT/math'",

        'post_test_attended_social': "post_test->>'ABLN_ASSESSMENT/attended_social'",
        'post_test_modality_social': "post_test->>'ABLN_ASSESSMENT/modality_social'",
        'post_test_social_emotional': "post_test->>'ABLN_ASSESSMENT/social_emotional'",
    })

    rows = qs.order_by('id').values_list(
        'id',
        'new_registry',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'rims_case_number',
        'source_of_identification',
        'source_of_identification_specify',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'main_caregiver_nationality_other',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'participation',
        'barriers_single',
        'barriers_other',
        'test_done',
        'round_complete',
        'basic_stationery',
        'pss_kit',
        'learning_result',
        'learning_result_other',
        'akelius_program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_number',
        'pss_session_modality',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_number',
        'covid_session_modality',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_number',
        'followup_session_modality',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'pre_test_attended_arabic',
        'pre_test_modality_arabic',
        'pre_test_arabic',
        'pre_test_attended_artistic',
        'pre_test_modality_artistic',
        'pre_test_artistic',
        'pre_test_attended_math',
        'pre_test_modality_math',
        'pre_test_math',
        'pre_test_attended_social',
        'pre_test_modality_social',
        'pre_test_social_emotional',
        'post_test_attended_arabic',
        'post_test_modality_arabic',
        'post_test_arabic',
        'post_test_attended_artistic',
        'post_test_modality_artistic',
        'post_test_artistic',
        'post_test_attended_math',
        'post_test_modality_math',
        'post_test_math',
        'post_test_attended_social',
        'post_test_modality_social',
        'post_test_social_emotional',
        'owner__username',
        'modified_by__username',
    )
    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            ws.write(row_num_student, col_num, row[col_num], font_style)
    # FC
    wsFC = wb_student.add_sheet('FC')
    row_num_fc = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    column_header_fc = [
        'enrollment id',
        'Partner',
        'CLM Round',
        'First time registered?',
        'Registration level',
        'Student first name',
        'Student father name',
        'Student last name',
        'unique number',
        'Gender',
        'Student Nationality',
        'Does the child participate in work?',
        'Does the child have any disability or special need?',
        'fc type',
        'facilitator name',
        'subject taught',
        'date of monitoring',
        'targeted competencies',
        'activities reported',
        'activities reported other',
        'share expectations',
        'share expectations no reason',
        'share expectations other reason',
        'materials needed available',
        'attend lesson',
        'child interact teacher',
        'child interact friends',
        'child clear responses',
        'child ask questions',
        'child acquire competency',
        'child show improvement',
        'child expected work independently',
        'work independently evaluation',
        'complete printed package',
        'sessions participated',
        'not participating reason',
        'E recharge card provided',
        'action to taken',
        'action to taken specify',
        'child needs pss',
        'child cant access resources',
        'homework after lesson',
        'parents supporting student',
        'completed tasks',
        'meet objectives',
        'meet objectives verified',
        'objectives verified specify',
        'additional notes'
    ]

    for col_num in range(len(column_header_fc)):
        wsFC.write(row_num_fc, col_num, column_header_fc[col_num], font_style)
    font_style = xlwt.XFStyle()

    rows_fc = queryset_fc.values_list(

        'enrollment_id',
        'enrollment__partner__name',
        'enrollment__round__name',
        'enrollment__new_registry',
        'enrollment__registration_level',
        'enrollment__student__first_name',
        'enrollment__student__father_name',
        'enrollment__student__last_name',
        'enrollment__student__number',
        'enrollment__student__sex',
        'enrollment__student__nationality__name',
        'enrollment__disability__name_en',
        'enrollment__have_labour_single_selection',
        'fc_type',
        'facilitator_name',
        'subject_taught',
        'date_of_monitoring',
        'targeted_competencies',
        'activities_reported',
        'activities_reported_other',
        'share_expectations',
        'share_expectations_no_reason',
        'share_expectations_other_reason',
        'materials_needed_available',
        'attend_lesson',
        'child_interact_teacher',
        'child_interact_friends',
        'child_clear_responses',
        'child_ask_questions',
        'child_acquire_competency',
        'child_show_improvement',
        'child_expected_work_independently',
        'work_independently_evaluation',
        'complete_printed_package',
        'sessions_participated',
        'not_participating_reason',
        'e_recharge_card_provided',
        'action_to_taken',
        'action_to_taken_specify',
        'child_needs_pss',
        'child_cant_access_resources',
        'homework_after_lesson',
        'parents_supporting_student',
        'completed_tasks',
        'meet_objectives',
        'meet_objectives_verified',
        'objectives_verified_specify',
        'additional_notes'

    )

    for row in rows_fc:
        row_num_fc += 1
        for col_num in range(len(row)):
            wsFC.write(row_num_fc, col_num, row[col_num], font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ABLN.xls"'

    return response


def cbece_build_xls_extraction(queryset_students, queryset_fc):
    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        'enrollment_id',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'first attendance date',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'RIMS  case number',
        'Source of Identification',
        'Source of Identification Specify',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'test_done',
        'round_complete',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Did the child use Akelius program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session number',
        'pss session modality',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'pre test attended arabic',
        'pre test modality arabic',
        'pre test arabic',
        'pre test attended english',
        'pre test modality english',
        'pre test english',
        'pre test attended psychomotor',
        'pre test modality psychomotor',
        'pre test psychomotor',
        'pre test attended artistic',
        'pre test modality artistic',
        'pre test artistic',
        'pre test attended math',
        'pre test modality math',
        'pre test math',
        'pre test attended social',
        'pre test modality social',
        'pre test social emotional',
        'pre test attended science',
        'pre test modality science',
        'pre test science',
        'post test attended arabic',
        'post test modality arabic',
        'post test arabic',
        'post test attended english',
        'post test modality english',
        'post test english',
        'post test attended psychomotor',
        'post test modality psychomotor',
        'post test psychomotor',
        'post test attended artistic',
        'post test modality artistic',
        'post test artistic',
        'post test attended math',
        'post test modality math',
        'post test math',
        'post test attended social',
        'post test modality social',
        'post test social emotional',
        'post test attended science',
        'post test modality science',
        'post test science',
        'owner',
        'modified_by',
    ]
    for col_num in range(len(columns)):
        ws.write(row_num_student, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    qs = queryset_students.extra(select={

        'pre_test_attended_arabic': "pre_test->>'CBECE_ASSESSMENT/attended_arabic'",
        'pre_test_modality_arabic': "pre_test->>'CBECE_ASSESSMENT/modality_arabic'",
        'pre_test_arabic': "pre_test->>'CBECE_ASSESSMENT/arabic'",

        'pre_test_attended_english': "pre_test->>'CBECE_ASSESSMENT/attended_english'",
        'pre_test_modality_english': "pre_test->>'CBECE_ASSESSMENT/modality_english'",
        'pre_test_english': "pre_test->>'CBECE_ASSESSMENT/english'",

        'pre_test_attended_psychomotor': "pre_test->>'CBECE_ASSESSMENT/attended_psychomotor'",
        'pre_test_modality_psychomotor': "pre_test->>'CBECE_ASSESSMENT/modality_psychomotor'",
        'pre_test_psychomotor': "pre_test->>'CBECE_ASSESSMENT/psychomotor'",

        'pre_test_attended_math': "pre_test->>'CBECE_ASSESSMENT/attended_math'",
        'pre_test_modality_math': "pre_test->>'CBECE_ASSESSMENT/modality_math'",
        'pre_test_math': "pre_test->>'CBECE_ASSESSMENT/math'",

        'pre_test_attended_social': "pre_test->>'CBECE_ASSESSMENT/attended_social'",
        'pre_test_modality_social': "pre_test->>'CBECE_ASSESSMENT/modality_social'",
        'pre_test_social_emotional': "pre_test->>'CBECE_ASSESSMENT/social_emotional'",

        'pre_test_attended_science': "pre_test->>'CBECE_ASSESSMENT/attended_science'",
        'pre_test_modality_science': "pre_test->>'CBECE_ASSESSMENT/modality_science'",
        'pre_test_science': "pre_test->>'CBECE_ASSESSMENT/science'",

        'pre_test_attended_artistic': "pre_test->>'CBECE_ASSESSMENT/attended_artistic'",
        'pre_test_modality_artistic': "pre_test->>'CBECE_ASSESSMENT/modality_artistic'",
        'pre_test_artistic': "pre_test->>'CBECE_ASSESSMENT/artistic'",

        'post_test_attended_arabic': "post_test->>'CBECE_ASSESSMENT/attended_arabic'",
        'post_test_modality_arabic': "post_test->>'CBECE_ASSESSMENT/modality_arabic'",
        'post_test_arabic': "post_test->>'CBECE_ASSESSMENT/arabic'",

        'post_test_attended_english': "post_test->>'CBECE_ASSESSMENT/attended_english'",
        'post_test_modality_english': "post_test->>'CBECE_ASSESSMENT/modality_english'",
        'post_test_english': "post_test->>'CBECE_ASSESSMENT/english'",

        'post_test_attended_psychomotor': "post_test->>'CBECE_ASSESSMENT/attended_psychomotor'",
        'post_test_modality_psychomotor': "post_test->>'CBECE_ASSESSMENT/modality_psychomotor'",
        'post_test_psychomotor': "post_test->>'CBECE_ASSESSMENT/psychomotor'",

        'post_test_attended_math': "post_test->>'CBECE_ASSESSMENT/attended_math'",
        'post_test_modality_math': "post_test->>'CBECE_ASSESSMENT/modality_math'",
        'post_test_math': "post_test->>'CBECE_ASSESSMENT/math'",

        'post_test_attended_social': "post_test->>'CBECE_ASSESSMENT/attended_social'",
        'post_test_modality_social': "post_test->>'CBECE_ASSESSMENT/modality_social'",
        'post_test_social_emotional': "post_test->>'CBECE_ASSESSMENT/social_emotional'",

        'post_test_attended_science': "post_test->>'CBECE_ASSESSMENT/attended_science'",
        'post_test_modality_science': "post_test->>'CBECE_ASSESSMENT/modality_science'",
        'post_test_science': "post_test->>'CBECE_ASSESSMENT/science'",

        'post_test_attended_artistic': "post_test->>'CBECE_ASSESSMENT/attended_artistic'",
        'post_test_modality_artistic': "post_test->>'CBECE_ASSESSMENT/modality_artistic'",
        'post_test_artistic': "post_test->>'CBECE_ASSESSMENT/artistic'",
    })

    rows = qs.order_by('id').values_list(
        'id',
        'new_registry',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'rims_case_number',
        'source_of_identification',
        'source_of_identification_specify',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'participation',
        'barriers_single',
        'barriers_other',
        'test_done',
        'round_complete',
        'basic_stationery',
        'pss_kit',
        'learning_result',
        'learning_result_other',
        'akelius_program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_number',
        'pss_session_modality',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_number',
        'covid_session_modality',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_number',
        'followup_session_modality',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'pre_test_attended_arabic',
        'pre_test_modality_arabic',
        'pre_test_arabic',
        'pre_test_attended_english',
        'pre_test_modality_english',
        'pre_test_english',
        'pre_test_attended_psychomotor',
        'pre_test_modality_psychomotor',
        'pre_test_psychomotor',
        'pre_test_attended_artistic',
        'pre_test_modality_artistic',
        'pre_test_artistic',
        'pre_test_attended_math',
        'pre_test_modality_math',
        'pre_test_math',
        'pre_test_attended_social',
        'pre_test_modality_social',
        'pre_test_social_emotional',
        'pre_test_attended_science',
        'pre_test_modality_science',
        'pre_test_science',
        'post_test_attended_arabic',
        'post_test_modality_arabic',
        'post_test_arabic',
        'post_test_attended_english',
        'post_test_modality_english',
        'post_test_english',
        'post_test_attended_psychomotor',
        'post_test_modality_psychomotor',
        'post_test_psychomotor',
        'post_test_attended_artistic',
        'post_test_modality_artistic',
        'post_test_artistic',
        'post_test_attended_math',
        'post_test_modality_math',
        'post_test_math',
        'post_test_attended_social',
        'post_test_modality_social',
        'post_test_social_emotional',
        'post_test_attended_science',
        'post_test_modality_science',
        'post_test_science',
        'owner__username',
        'modified_by__username',
    )
    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            ws.write(row_num_student, col_num, row[col_num], font_style)
    # FC
    wsFC = wb_student.add_sheet('FC')
    row_num_fc = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    column_header_fc = [
        'enrollment id',
        'Partner',
        'CLM Round',
        'First time registered?',
        'Registration level',
        'Student first name',
        'Student father name',
        'Student last name',
        'unique number',
        'Gender',
        'Student Nationality',
        'Does the child participate in work?',
        'Does the child have any disability or special need?',
        'fc type',
        'facilitator name',
        'subject taught',
        'date of monitoring',
        'targeted competencies',
        'activities reported',
        'activities reported other',
        'share expectations',
        'share expectations no reason',
        'share expectations other reason',
        'materials needed available',
        'attend lesson',
        'child interact teacher',
        'child interact friends',
        'child clear responses',
        'child ask questions',
        'child acquire competency',
        'child show improvement',
        'child expected work independently',
        'work independently evaluation',
        'complete printed package',
        'sessions participated',
        'not participating reason',
        'E recharge card provided',
        'action to taken',
        'action to taken specify',
        'child needs pss',
        'child cant access resources',
        'homework after lesson',
        'parents supporting student',
        'completed tasks',
        'meet objectives',
        'meet objectives verified',
        'objectives verified specify',
        'additional notes'
    ]

    for col_num in range(len(column_header_fc)):
        wsFC.write(row_num_fc, col_num, column_header_fc[col_num], font_style)
    font_style = xlwt.XFStyle()

    rows_fc = queryset_fc.values_list(

        'enrollment_id',
        'enrollment__partner__name',
        'enrollment__round__name',
        'enrollment__new_registry',
        'enrollment__registration_level',
        'enrollment__student__first_name',
        'enrollment__student__father_name',
        'enrollment__student__last_name',
        'enrollment__student__number',
        'enrollment__student__sex',
        'enrollment__student__nationality__name',
        'enrollment__disability__name_en',
        'enrollment__have_labour_single_selection',
        'fc_type',
        'facilitator_name',
        'subject_taught',
        'date_of_monitoring',
        'targeted_competencies',
        'activities_reported',
        'activities_reported_other',
        'share_expectations',
        'share_expectations_no_reason',
        'share_expectations_other_reason',
        'materials_needed_available',
        'attend_lesson',
        'child_interact_teacher',
        'child_interact_friends',
        'child_clear_responses',
        'child_ask_questions',
        'child_acquire_competency',
        'child_show_improvement',
        'child_expected_work_independently',
        'work_independently_evaluation',
        'complete_printed_package',
        'sessions_participated',
        'not_participating_reason',
        'e_recharge_card_provided',
        'action_to_taken',
        'action_to_taken_specify',
        'child_needs_pss',
        'child_cant_access_resources',
        'homework_after_lesson',
        'parents_supporting_student',
        'completed_tasks',
        'meet_objectives',
        'meet_objectives_verified',
        'objectives_verified_specify',
        'additional_notes'

    )

    for row in rows_fc:
        row_num_fc += 1
        for col_num in range(len(row)):
            wsFC.write(row_num_fc, col_num, row[col_num], font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="CBECE.xls"'

    return response


# def rs_build_xls_extraction(queryset_students):
def rs_build_xls_extraction(queryset_students, queryset_fc):
    buffer = io.BytesIO()

    wb = Workbook()

    ws_student = wb['Sheet']
    ws_student.title = 'Student'

    # Sheet header, first row
    row_num_student = 0

    font_style = Font(bold=True)

    columns = [
        'enrollment_id',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Grade of registration',
        'Student Address',
        'Registration level',
        'first attendance date',
        'School of Enrollment',
        'shift',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'What was the child education level when first joining formal education in lebanon',
        'From where did the child first come to join  FE',
        'Education status',
        'Miss school?',
        'Miss school date',
        'Internal number',
        'RIMS  case number',
        'Source of Identification',
        'Source of Identification Specify',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'test_done',
        'round_complete',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Did the child use Akelius program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session number',
        'pss session modality',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'pre test attended arabic',
        'pre test modality arabic',
        'pre test arabic',
        'pre test attended english',
        'pre test modality english',
        'pre test english',
        'pre test attended math',
        'pre test modality math',
        'pre test math',
        'pre test attended science',
        'pre test modality science',
        'pre test science',
        'pre test attended biology',
        'pre test modality biology',
        'pre test biology',
        'pre test attended physics',
        'pre test modality physics',
        'pre test physics',
        'pre test attended chemistry',
        'pre test modality chemistry',
        'pre test chemistry',
        'post test attended arabic',
        'post test modality arabic',
        'post test arabic',
        'post test attended english',
        'post test modality english',
        'post test english',
        'post test attended math',
        'post test modality math',
        'post test math',
        'post test attended science',
        'post test modality science',
        'post test science',
        'post test attended biology',
        'post test modality biology',
        'post test biology',
        'post test attended physics',
        'post test modality physics',
        'post test physics',
        'post test attended chemistry',
        'post test modality chemistry',
        'post test chemistry',
        'owner',
        'modified_by',
    ]

    for col_num in range(len(columns)):
        currentCell = ws_student.cell(row=row_num_student + 1, column=col_num + 1, value=columns[col_num])
        currentCell.font = font_style

    qs = queryset_students.extra(select={

        'pre_test_attended_arabic': "pre_test->>'RS_ASSESSMENT/attended_arabic'",
        'pre_test_modality_arabic': "pre_test->>'RS_ASSESSMENT/modality_arabic'",
        'pre_test_arabic': "pre_test->>'RS_ASSESSMENT/arabic'",

        'pre_test_attended_english': "pre_test->>'RS_ASSESSMENT/attended_english'",
        'pre_test_modality_english': "pre_test->>'RS_ASSESSMENT/modality_english'",
        'pre_test_english': "pre_test->>'RS_ASSESSMENT/english'",

        'pre_test_attended_math': "pre_test->>'RS_ASSESSMENT/attended_math'",
        'pre_test_modality_math': "pre_test->>'RS_ASSESSMENT/modality_math'",
        'pre_test_math': "pre_test->>'RS_ASSESSMENT/math'",

        'pre_test_attended_science': "pre_test->>'RS_ASSESSMENT/attended_science'",
        'pre_test_modality_science': "pre_test->>'RS_ASSESSMENT/modality_science'",
        'pre_test_science': "pre_test->>'RS_ASSESSMENT/science'",

        'pre_test_attended_biology': "pre_test->>'RS_ASSESSMENT/attended_biology'",
        'pre_test_modality_biology': "pre_test->>'RS_ASSESSMENT/modality_biology'",
        'pre_test_biology': "pre_test->>'RS_ASSESSMENT/biology'",

        'pre_test_attended_physics': "pre_test->>'RS_ASSESSMENT/attended_physics'",
        'pre_test_modality_physics': "pre_test->>'RS_ASSESSMENT/modality_physics'",
        'pre_test_physics': "pre_test->>'RS_ASSESSMENT/physics'",

        'pre_test_attended_chemistry': "pre_test->>'RS_ASSESSMENT/attended_chemistry'",
        'pre_test_modality_chemistry': "pre_test->>'RS_ASSESSMENT/modality_chemistry'",
        'pre_test_chemistry': "pre_test->>'RS_ASSESSMENT/chemistry'",

        'post_test_attended_arabic': "post_test->>'RS_ASSESSMENT/attended_arabic'",
        'post_test_modality_arabic': "post_test->>'RS_ASSESSMENT/modality_arabic'",
        'post_test_arabic': "post_test->>'RS_ASSESSMENT/arabic'",

        'post_test_attended_english': "post_test->>'RS_ASSESSMENT/attended_english'",
        'post_test_modality_english': "post_test->>'RS_ASSESSMENT/modality_english'",
        'post_test_english': "post_test->>'RS_ASSESSMENT/english'",

        'post_test_attended_math': "post_test->>'RS_ASSESSMENT/attended_math'",
        'post_test_modality_math': "post_test->>'RS_ASSESSMENT/modality_math'",
        'post_test_math': "post_test->>'RS_ASSESSMENT/math'",

        'post_test_attended_science': "post_test->>'RS_ASSESSMENT/attended_science'",
        'post_test_modality_science': "post_test->>'RS_ASSESSMENT/modality_science'",
        'post_test_science': "post_test->>'RS_ASSESSMENT/science'",

        'post_test_attended_biology': "post_test->>'RS_ASSESSMENT/attended_biology'",
        'post_test_modality_biology': "post_test->>'RS_ASSESSMENT/modality_biology'",
        'post_test_biology': "post_test->>'RS_ASSESSMENT/biology'",

        'post_test_attended_physics': "post_test->>'RS_ASSESSMENT/attended_physics'",
        'post_test_modality_physics': "post_test->>'RS_ASSESSMENT/modality_physics'",
        'post_test_physics': "post_test->>'RS_ASSESSMENT/physics'",

        'post_test_attended_chemistry': "post_test->>'RS_ASSESSMENT/attended_chemistry'",
        'post_test_modality_chemistry': "post_test->>'RS_ASSESSMENT/modality_chemistry'",
        'post_test_chemistry': "post_test->>'RS_ASSESSMENT/chemistry'",

    })

    rows = qs.order_by('id').values_list(
        'id',
        'new_registry',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'grade_registration',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'registered_in_school',
        'shift',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'grade_level',
        'source_join_fe',
        'education_status',
        'miss_school',
        'miss_school_date',
        'internal_number',
        'rims_case_number',
        'source_of_identification',
        'source_of_identification_specify',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'participation',
        'barriers_single',
        'barriers_other',
        'test_done',
        'round_complete',
        'basic_stationery',
        'pss_kit',
        'learning_result',
        'learning_result_other',
        'akelius_program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_number',
        'pss_session_modality',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_number',
        'covid_session_modality',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_number',
        'followup_session_modality',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'pre_test_attended_arabic',
        'pre_test_modality_arabic',
        'pre_test_arabic',
        'pre_test_attended_english',
        'pre_test_modality_english',
        'pre_test_english',
        'pre_test_attended_math',
        'pre_test_modality_math',
        'pre_test_math',
        'pre_test_attended_science',
        'pre_test_modality_science',
        'pre_test_science',
        'pre_test_attended_biology',
        'pre_test_modality_biology',
        'pre_test_biology',
        'pre_test_attended_physics',
        'pre_test_modality_physics',
        'pre_test_physics',
        'pre_test_attended_chemistry',
        'pre_test_modality_chemistry',
        'pre_test_chemistry',
        'post_test_attended_arabic',
        'post_test_modality_arabic',
        'post_test_arabic',
        'post_test_attended_english',
        'post_test_modality_english',
        'post_test_english',
        'post_test_attended_math',
        'post_test_modality_math',
        'post_test_math',
        'post_test_attended_science',
        'post_test_modality_science',
        'post_test_science',
        'post_test_attended_biology',
        'post_test_modality_biology',
        'post_test_biology',
        'post_test_attended_physics',
        'post_test_modality_physics',
        'post_test_physics',
        'post_test_attended_chemistry',
        'post_test_modality_chemistry',
        'post_test_chemistry',
        'owner__username',
        'modified_by__username',
    )
    # [:5000]

    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            cellValue = row[col_num]
            cellTextValue = cellValue
            if isinstance(cellValue, str):
                cellTextValue = cellValue.encode("utf8")
            if type(cellValue) in [int]:
                cellTextValue = str(cellValue).encode("utf8")
            if type(cellValue) in [unicode, date]:
                cellTextValue = cellValue
            if type(cellValue) in [list]:
                cellTextValue = listToString(cellValue)
            currentCell = ws_student.cell(row=row_num_student + 1, column=col_num + 1, value=cellTextValue)
            currentCell.font = font_style
    #get partners
    rows_partners = qs.distinct('partner_id').order_by('partner_id').values_list(
        'partner__id',
        'partner__name',
    )
    for partner in rows_partners:
        partner_id = partner[0]
        partner_name = str(partner[1]).encode("utf8")

        ws_fc = wb.create_sheet('FC - ' + partner_name)
        # Sheet header, first row
        row_num_fc = 0
        columns_fc = [
                'enrollment id',
                'Partner',
                'CLM Round',
                'First time registered?',
                'Registration level',
                'Student first name',
                'Student father name',
                'Student last name',
                'unique number',
                'Gender',
                'Student Nationality',
                'Does the child participate in work?',
                'Does the child have any disability or special need?',
                'fc type',
                'facilitator name',
                'subject taught',
                'date of monitoring',
                'targeted competencies',
                'activities reported',
                'activities reported other',
                'share expectations',
                'share expectations no reason',
                'share expectations other reason',
                'materials needed available',
                'attend lesson',
                'child interact teacher',
                'child interact friends',
                'child clear responses',
                'child ask questions',
                'child acquire competency',
                'child show improvement',
                'child expected work independently',
                'work independently evaluation',
                'complete printed package',
                'sessions participated',
                'not participating reason',
                'E recharge card provided',
                'action to taken',
                'action to taken specify',
                'child needs pss',
                'child cant access resources',
                'homework after lesson',
                'parents supporting student',
                'completed tasks',
                'meet objectives',
                'meet objectives verified',
                'objectives verified specify',
                'additional notes'
            ]

        for col_num in range(len(columns_fc)):
            currentCell = ws_fc.cell(row=row_num_fc + 1, column=col_num + 1, value=columns_fc[col_num])
            currentCell.font = font_style

        rows_fc = queryset_fc.filter(enrollment__partner=partner_id).values_list(
            'enrollment_id',
            'enrollment__partner__name',
            'enrollment__round__name',
            'enrollment__new_registry',
            'enrollment__registration_level',
            'enrollment__student__first_name',
            'enrollment__student__father_name',
            'enrollment__student__last_name',
            'enrollment__student__number',
            'enrollment__student__sex',
            'enrollment__student__nationality__name',
            'enrollment__disability__name_en',
            'enrollment__have_labour_single_selection',
            'fc_type',
            'facilitator_name',
            'subject_taught',
            'date_of_monitoring',
            'targeted_competencies',
            'activities_reported',
            'activities_reported_other',
            'share_expectations',
            'share_expectations_no_reason',
            'share_expectations_other_reason',
            'materials_needed_available',
            'attend_lesson',
            'child_interact_teacher',
            'child_interact_friends',
            'child_clear_responses',
            'child_ask_questions',
            'child_acquire_competency',
            'child_show_improvement',
            'child_expected_work_independently',
            'work_independently_evaluation',
            'complete_printed_package',
            'sessions_participated',
            'not_participating_reason',
            'e_recharge_card_provided',
            'action_to_taken',
            'action_to_taken_specify',
            'child_needs_pss',
            'child_cant_access_resources',
            'homework_after_lesson',
            'parents_supporting_student',
            'completed_tasks',
            'meet_objectives',
            'meet_objectives_verified',
            'objectives_verified_specify',
            'additional_notes'
        )
        # [:5000]

        for row in rows_fc:
            row_num_fc += 1
            for col_num in range(len(row)):
                cellValue = row[col_num]
                cellTextValue = cellValue
                if isinstance(cellValue, str):
                    cellTextValue = cellValue.encode("utf8")
                if type(cellValue) in [int]:
                    cellTextValue = str(cellValue).encode("utf8")
                if type(cellValue) in [unicode, date]:
                    cellTextValue = cellValue
                if type(cellValue) in [list]:
                    cellTextValue = listToString(cellValue)
                currentCell = ws_fc.cell(row=row_num_fc + 1, column=col_num + 1, value=cellTextValue)
                currentCell.font = font_style


    # Save the file
    wb.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="RS.xlsx"'

    return response


def outreach_build_xls_extraction(queryset_students):
    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        'enrollment_id',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'First attendance date',
        'Unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'Source of Identification',
        'Source of Identification Specify',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'main caregiver nationality Other ',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'owner',
        'modified_by',
    ]

    for col_num in range(len(columns)):
        ws.write(row_num_student, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = queryset_students.order_by('id').values_list(
        'id',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'source_of_identification',
        'source_of_identification_specify',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'main_caregiver_nationality_other',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'owner__username',
        'modified_by__username',
    )
    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            ws.write(row_num_student, col_num, row[col_num], font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Outreach.xls"'

    return response


def bridging_build_xls_extraction(queryset_students):
    from student_registration.schools.models import CLMRound
    current_round = CLMRound.objects.get(current_round_bridging=True)
    round_start_date = current_round.start_date_bridging
    round_end_date = current_round.end_date_bridging

    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns_titles = [
        'enrollment ID',
        'Student ID',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Round start date',
        'Governorate',
        'District',
        'Cadaster',
        'School',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'first attendance date',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'age',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'Source of Identification',
        'Source of Identification Specify',
        'RIMS  case number',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'other caregiver relationship',
        'main caregiver nationality',
        'main caregiver nationality Other ',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'pre test arabic',
        'pre test english',
        'pre test math',
        'post test arabic',
        'post test english',
        'post test math',
        'The language supported in the program',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'Was the community Liaison at school level involved in follow up on child absence or drop out?',
        'specify',
        'Post test has been done',
        'round_complete',
        'Is the child receiving social assistance?',
        'Is the Child Receiving Transportation Support?',
        'Is the Child Using a digital platform (Akelius or  Learning Passport)',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'child_received_printout',
        'child_received_internet',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session modality',
        'pss session number',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please indicate modality',
        'Please enter the number of sessions',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please indicate modality',
        'Please enter the number of sessions',
        'Parent who attended the parents meeting',
        'Please specify',
        'Have the child caretakers been contacted by the School Community Laison',
        'Please specify what has been discussed',
        'owner',
        'modified_by',
        'Attendance Days',
        'Absent Days',
        'Period Out School'
    ]

    for col_num in range(len(columns_titles)):
        ws.write(row_num_student, col_num, columns_titles[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    columns = [
        'enrollment ID',
        'Student ID',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Round start date',
        'Governorate',
        'District',
        'Cadaster',
        'School',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'first attendance date',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Gender',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'Source of Identification',
        'Source of Identification Specify',
        'RIMS  case number',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'other caregiver relationship',
        'main caregiver nationality',
        'main caregiver nationality Other ',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'pre test arabic',
        'pre test english',
        'pre test math',
        'post test arabic',
        'post test english',
        'post test math',
        'The language supported in the program',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'Was the community Liaison at school level involved in follow up on child absence or drop out?',
        'specify',
        'Post test has been done',
        'round_complete',
        'Is the child receiving social assistance?',
        'Is the Child Receiving Transportation Support?',
        'Is the Child Using a digital platform (Akelius or  Learning Passport)',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'child_received_printout',
        'child_received_internet',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session modality',
        'pss session number',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please indicate modality',
        'Please enter the number of sessions',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please indicate modality',
        'Please enter the number of sessions',
        'Parent who attended the parents meeting',
        'Please specify',
        'Have the child caretakers been contacted by the School Community Laison',
        'Please specify what has been discussed',
        'owner',
        'modified_by',
        'Attendance Days',
        'Absent Days',
        'Period Out School'
    ]

    qs = queryset_students.extra(select={
        'pre_test_arabic': "pre_test->>'Bridging_ASSESSMENT/arabic'",
        'pre_test_english': "pre_test->>'Bridging_ASSESSMENT/english'",
        'pre_test_math': "pre_test->>'Bridging_ASSESSMENT/math'",
        'post_test_arabic': "post_test->>'Bridging_ASSESSMENT/arabic'",
        'post_test_english': "post_test->>'Bridging_ASSESSMENT/english'",
        'post_test_math': "post_test->>'Bridging_ASSESSMENT/math'",
    })

    rows = qs.order_by('id').values_list(
        'id',
        'student_id',
        'new_registry',
        'partner__name',
        'round__name',
        'round_start_date',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'school__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'source_of_identification',
        'source_of_identification_specify',
        'rims_case_number',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'other_caregiver_relationship',
        'main_caregiver_nationality__name',
        'main_caregiver_nationality_other',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'pre_test_arabic',
        'pre_test_english',
        'pre_test_math',
        'post_test_arabic',
        'post_test_english',
        'post_test_math',
        'language',
        'participation',
        'barriers_single',
        'barriers_other',
        'community_Liaison_follow_up',
        'community_liaison_specify',
        'test_done',
        'round_complete',
        'receiving_social_assistance',
        'receiving_transportation_support',
        'using_digital_platform',
        'basic_stationery',
        'pss_kit',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_internet',
        'learning_result',
        'learning_result_other',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_modality',
        'pss_session_number',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_modality',
        'covid_session_number',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_modality',
        'followup_session_number',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'school_contacted_caretaker',
        'discussion_topic',
        'owner__username',
        'modified_by__username',
    )

    for row in rows:
        row_num_student += 1
        year = row[columns.index("Birthday - year")]
        month = row[columns.index("Birthday - month")]
        day = row[columns.index("Birthday - day")]
        age = Person.get_age(year, month, day)
        student_id = row[columns.index("Student ID")]
        attendance_days = Bridging.get_attendance_days(student_id, round_start_date, round_end_date)
        absent_days = Bridging.get_absent_days(student_id, round_start_date, round_end_date)
        miss_school_date = row[columns.index("Miss school date")]
        period_out_school = Bridging.get_period_out_school(student_id, miss_school_date, round_start_date)

        for col_num in range(len(row)):
            if col_num == columns.index("Birthday - year"):
                ws.write(row_num_student, col_num, row[col_num], font_style)
                ws.write(row_num_student, col_num + 1, age, font_style)
            elif col_num > columns.index("Birthday - year"):
                ws.write(row_num_student, col_num + 1, row[col_num], font_style)
            else:
                ws.write(row_num_student, col_num, row[col_num], font_style)

        ws.write(row_num_student, col_num + 2, attendance_days, font_style)
        ws.write(row_num_student, col_num + 3, absent_days, font_style)
        ws.write(row_num_student, col_num + 4, period_out_school, font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Bridging.xls"'

    return response


def listToString(s):
    # initialize an empty string
    str1 = ""

    # traverse in the string
    for ele in s:
        if str1 == "":
            str1 += ele
        else:
            str1 += "," + ele

        # return string
    return str1


def build_xls_extraction_horizental(queryset_students, queryset_fc):
    buffer = io.BytesIO()

    # Personnel
    wb_student = xlwt.Workbook(encoding='utf-8', style_compression=2)

    ws = wb_student.add_sheet('Student')

    # Sheet header, first row
    row_num_student = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        'enrollment_id',
        'First time registered?',
        'Partner',
        'CLM Round',
        'Governorate',
        'District',
        'Cadaster',
        'Location',
        'Center',
        'The language supported in the program',
        'Student Address',
        'Registration level',
        'first attendance date',
        'ID number',
        'unique number',
        'First name',
        'Father name',
        'Last name',
        'Mother fullname',
        'Sex',
        'Student Nationality',
        'Student Nationality Specify',
        'Birthday - day',
        'Birthday - month',
        'Birthday - year',
        'P-Code If a child lives in a tent / Brax in a random camp',
        'Does the child have any disability or special need?',
        'Education status',
        'Miss school date',
        'Internal number',
        'RIMS  case number',
        'Source of Identification',
        'Source of Identification Specify',
        'Source of Transportation',
        'What is the educational level of the mother?',
        'What is the educational level of the father?',
        'Phone number',
        'Phone number confirm',
        'phone owner',
        'Second Phone number',
        'Second Phone number confirm',
        'Second phone owner',
        'Main Caregiver',
        'main caregiver nationality',
        'other caregiver relationship',
        'main caregiver nationality Other ',
        'Caretaker first name',
        'Caretaker middle name',
        'Caretaker last name',
        'Caretaker mother name',
        'caretaker birthday year',
        'caretaker birthday month',
        'caretaker birthday day',
        'ID Type',
        'UNHCR case number',
        'UNHCR case number confirm',
        'Parent individual ID',
        'Parent individual ID confirm',
        'Child individual ID',
        'Child individual ID confirm',
        'UNHCR recorded barcode',
        'UNHCR recorded barcode confirm',
        'Parent Lebanese ID number',
        'Parent Lebanese ID number confirm',
        'Child Lebanese ID number',
        'Child Lebanese ID number confirm',
        'Parent Syrian ID number',
        'Parent Syrian ID number confirm',
        'Child Syrian ID number',
        'Child Syrian ID number confirm',
        'Parent Palestinian ID number',
        'Parent Palestinian ID number confirm',
        'Child Palestinian ID number',
        'Child Palestinian ID number confirm',
        'ID number of the Caretaker',
        'ID number of the Caretaker confirm',
        'ID number of the child',
        'ID number of the child confirm',
        'What is the family status of the child?',
        'Does the child have children?',
        'Child number of children',
        'Does the child participate in work?',
        'What is the type of work?',
        'Please specify',
        'How many hours does this child work in a day?',
        'Child weekly income',
        'Level of participation / Absence',
        'The main barriers affecting the daily attendance and performance of the child or drop out of',
        'Please specify',
        'test_done',
        'round_complete',
        'Did the child receive basic stationery?',
        'Did the child benefit from the PSS kit?',
        'Based on the overall score, what is the recommended learning path?',
        'Please specify',
        'Did the child use Akelius program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'Please enter the number phone calls',
        'Phone call Result of follow up',
        'Please enter the number of house visits',
        'House VisitvResult of follow up',
        'Please enter the number of family visit',
        'Familiy Visit Result of follow up',
        'Parent attended visits',
        'pss session attended',
        'pss session number',
        'pss session modality',
        'pss parent attended',
        'pss parent attended other',
        'Attended covid Session?',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'Attended followup Session',
        'Please enter the number of sessions',
        'Please indicate modality',
        'Parent who attended the parents meeting',
        'Please specify',
        'pre test attended arabic',
        'pre test modality arabic',
        'pre test arabic',
        'pre test attended english',
        'pre test modality english',
        'pre test english',
        'pre test attended psychomotor',
        'pre test modality psychomotor',
        'pre test psychomotor',
        'pre test attended artistic',
        'pre test modality artistic',
        'pre test artistic',
        'pre test attended math',
        'pre test modality math',
        'pre test math',
        'pre test attended social',
        'pre test modality social',
        'pre test social emotional',
        'post test attended arabic',
        'post test modality arabic',
        'post test arabic',
        'post test attended english',
        'post test modality english',
        'post test english',
        'post test attended psychomotor',
        'post test modality psychomotor',
        'post test psychomotor',
        'post test attended artistic',
        'post test modality artistic',
        'post test artistic',
        'post test attended math',
        'post test modality math',
        'post test math',
        'post test attended social',
        'post test modality social',
        'post test social emotional',
        'Was the child involved in remote learning?',
        'what other reasons for this child not being engaged?',
        'reasons not engaged other',
        'Does the family have reliable internet service in their area during remote learning?',
        'Did both girls and boys in the same family participate in the class and have access to the phone/',
        'Explain',
        'Frequency of Child Engagement in remote learning?',
        'How well did the child meet the learning outcomes?',
        'How do you rate the parents learning support provided to the child through this Remote',
        'Has the child directly been reached with awareness messaging on Covid-19 and prevention measures?',
        'How often?',
        'Has the parents directly been reached with awareness messaging on Covid-19 and prevention',
        'How often?',
        'Was any follow-up done to ensure messages were well received, understood and adopted?',
        'With who child and/or caregiver?',
        'Reason why not doing the Pre-test',
        'Reason why not doing the Post-test',
        'Student outreached?',
        'owner',
        'modified_by',
    ]

    for col_num in range(len(columns)):
        ws.write(row_num_student, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    qs = queryset_students.extra(select={

        'pre_test_attended_arabic': "pre_test->>'BLN_ASSESSMENT/attended_arabic'",
        'pre_test_modality_arabic': "pre_test->>'BLN_ASSESSMENT/modality_arabic'",
        'pre_test_arabic': "pre_test->>'BLN_ASSESSMENT/arabic'",

        'pre_test_attended_english': "pre_test->>'BLN_ASSESSMENT/attended_english'",
        'pre_test_modality_english': "pre_test->>'BLN_ASSESSMENT/modality_english'",
        'pre_test_english': "pre_test->>'BLN_ASSESSMENT/english'",

        'pre_test_attended_psychomotor': "pre_test->>'BLN_ASSESSMENT/attended_psychomotor'",
        'pre_test_modality_psychomotor': "pre_test->>'BLN_ASSESSMENT/modality_psychomotor'",
        'pre_test_psychomotor': "pre_test->>'BLN_ASSESSMENT/psychomotor'",

        'pre_test_attended_artistic': "pre_test->>'BLN_ASSESSMENT/attended_artistic'",
        'pre_test_modality_artistic': "pre_test->>'BLN_ASSESSMENT/modality_artistic'",
        'pre_test_artistic': "pre_test->>'BLN_ASSESSMENT/artistic'",

        'pre_test_attended_math': "pre_test->>'BLN_ASSESSMENT/attended_math'",
        'pre_test_modality_math': "pre_test->>'BLN_ASSESSMENT/modality_math'",
        'pre_test_math': "pre_test->>'BLN_ASSESSMENT/math'",

        'pre_test_attended_social': "pre_test->>'BLN_ASSESSMENT/attended_social'",
        'pre_test_modality_social': "pre_test->>'BLN_ASSESSMENT/modality_social'",
        'pre_test_social_emotional': "pre_test->>'BLN_ASSESSMENT/social_emotional'",

        'post_test_attended_arabic': "post_test->>'BLN_ASSESSMENT/attended_arabic'",
        'post_test_modality_arabic': "post_test->>'BLN_ASSESSMENT/modality_arabic'",
        'post_test_arabic': "post_test->>'BLN_ASSESSMENT/arabic'",

        'post_test_attended_english': "post_test->>'BLN_ASSESSMENT/attended_english'",
        'post_test_modality_english': "post_test->>'BLN_ASSESSMENT/modality_english'",
        'post_test_english': "post_test->>'BLN_ASSESSMENT/english'",

        'post_test_attended_psychomotor': "post_test->>'BLN_ASSESSMENT/attended_psychomotor'",
        'post_test_modality_psychomotor': "post_test->>'BLN_ASSESSMENT/modality_psychomotor'",
        'post_test_psychomotor': "post_test->>'BLN_ASSESSMENT/psychomotor'",

        'post_test_attended_artistic': "post_test->>'BLN_ASSESSMENT/attended_artistic'",
        'post_test_modality_artistic': "post_test->>'BLN_ASSESSMENT/modality_artistic'",
        'post_test_artistic': "post_test->>'BLN_ASSESSMENT/artistic'",

        'post_test_attended_math': "post_test->>'BLN_ASSESSMENT/attended_math'",
        'post_test_modality_math': "post_test->>'BLN_ASSESSMENT/modality_math'",
        'post_test_math': "post_test->>'BLN_ASSESSMENT/math'",

        'post_test_attended_social': "post_test->>'BLN_ASSESSMENT/attended_social'",
        'post_test_modality_social': "post_test->>'BLN_ASSESSMENT/modality_social'",
        'post_test_social_emotional': "post_test->>'BLN_ASSESSMENT/social_emotional'",
    })

    rows = qs.order_by('id').values_list(
        'id',
        'new_registry',
        'partner__name',
        'round__name',
        'governorate__name_en',
        'district__name_en',
        'cadaster__name_en',
        'location',
        'center__name',
        'language',
        'student__address',
        'registration_level',
        'first_attendance_date',
        'student__id_number',
        'student__number',
        'student__first_name',
        'student__father_name',
        'student__last_name',
        'student__mother_fullname',
        'student__sex',
        'student__nationality__name',
        'other_nationality',
        'student__birthday_day',
        'student__birthday_month',
        'student__birthday_year',
        'student__p_code',
        'disability__name_en',
        'education_status',
        'miss_school_date',
        'internal_number',
        'rims_case_number',
        'source_of_identification',
        'source_of_identification_specify',
        'source_of_transportation',
        'hh_educational_level__name',
        'father_educational_level__name',
        'phone_number',
        'phone_number_confirm',
        'phone_owner',
        'second_phone_number',
        'second_phone_number_confirm',
        'second_phone_owner',
        'main_caregiver',
        'main_caregiver_nationality__name',
        'other_caregiver_relationship',
        'main_caregiver_nationality_other',
        'caretaker_first_name',
        'caretaker_middle_name',
        'caretaker_last_name',
        'caretaker_mother_name',
        'caretaker_birthday_year',
        'caretaker_birthday_month',
        'caretaker_birthday_day',
        'id_type',
        'case_number',
        'case_number_confirm',
        'parent_individual_case_number',
        'parent_individual_case_number_confirm',
        'individual_case_number',
        'individual_case_number_confirm',
        'recorded_number',
        'recorded_number_confirm',
        'parent_national_number',
        'parent_national_number_confirm',
        'national_number',
        'national_number_confirm',
        'parent_syrian_national_number',
        'parent_syrian_national_number_confirm',
        'syrian_national_number',
        'syrian_national_number_confirm',
        'parent_sop_national_number',
        'parent_sop_national_number_confirm',
        'sop_national_number',
        'sop_national_number_confirm',
        'parent_other_number',
        'parent_other_number_confirm',
        'other_number',
        'other_number_confirm',
        'student__family_status',
        'student__have_children',
        'student_number_children',
        'have_labour_single_selection',
        'labours_single_selection',
        'labours_other_specify',
        'labour_hours',
        'labour_weekly_income',
        'participation',
        'barriers_single',
        'barriers_other',
        'test_done',
        'round_complete',
        'basic_stationery',
        'pss_kit',
        'learning_result',
        'learning_result_other',
        'akelius_program',
        'cp_referral',
        'referal_wash',
        'referal_health',
        'referal_other',
        'referal_other_specify',
        'child_received_books',
        'child_received_printout',
        'child_received_internet',
        'phone_call_number',
        'phone_call_follow_up_result',
        'house_visit_number',
        'house_visit_follow_up_result',
        'family_visit_number',
        'family_visit_follow_up_result',
        'parent_attended_visits',
        'pss_session_attended',
        'pss_session_number',
        'pss_session_modality',
        'pss_parent_attended',
        'pss_parent_attended_other',
        'covid_session_attended',
        'covid_session_number',
        'covid_session_modality',
        'covid_parent_attended',
        'covid_parent_attended_other',
        'followup_session_attended',
        'followup_session_number',
        'followup_session_modality',
        'followup_parent_attended',
        'followup_parent_attended_other',
        'pre_test_attended_arabic',
        'pre_test_modality_arabic',
        'pre_test_arabic',
        'pre_test_attended_english',
        'pre_test_modality_english',
        'pre_test_english',
        'pre_test_attended_psychomotor',
        'pre_test_modality_psychomotor',
        'pre_test_psychomotor',
        'pre_test_attended_artistic',
        'pre_test_modality_artistic',
        'pre_test_artistic',
        'pre_test_attended_math',
        'pre_test_modality_math',
        'pre_test_math',
        'pre_test_attended_social',
        'pre_test_modality_social',
        'pre_test_social_emotional',
        'post_test_attended_arabic',
        'post_test_modality_arabic',
        'post_test_arabic',
        'post_test_attended_english',
        'post_test_modality_english',
        'post_test_english',
        'post_test_attended_psychomotor',
        'post_test_modality_psychomotor',
        'post_test_psychomotor',
        'post_test_attended_artistic',
        'post_test_modality_artistic',
        'post_test_artistic',
        'post_test_attended_math',
        'post_test_modality_math',
        'post_test_math',
        'post_test_attended_social',
        'post_test_modality_social',
        'post_test_social_emotional',
        'remote_learning',
        'remote_learning_reasons_not_engaged',
        'reasons_not_engaged_other',
        'reliable_internet',
        'gender_participate',
        'gender_participate_explain',
        'remote_learning_engagement',
        'meet_learning_outcomes',
        'parent_learning_support_rate',
        'covid_message',
        'covid_message_how_often',
        'covid_parents_message',
        'covid_parents_message_how_often',
        'follow_up_done',
        'follow_up_done_with_who',
        'unsuccessful_pretest_reason',
        'unsuccessful_posttest_reason',
        'student_outreached',
        'owner__username',
        'modified_by__username',
        # 'created',
        # 'modified',
    )
    for row in rows:
        row_num_student += 1
        for col_num in range(len(row)):
            ws.write(row_num_student, col_num, row[col_num], font_style)

    # FC
    wsFC = wb_student.add_sheet('FC')
    row_num_fc = 0
    row_num_header_fc = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns_fc_student = [
        'enrollment id',
        'Student first name',
        'Student father name',
        'Student last name',
    ]
    for col_num in range(len(columns_fc_student)):
        wsFC.write(row_num_header_fc, col_num, columns_fc_student[col_num], font_style)

    font_style = xlwt.XFStyle()

    columns_fc = [
        'fc type',
        'facilitator name',
        'subject taught',
        'date of monitoring',
        'targeted competencies',
        'activities reported',
        'activities reported other',
        'share expectations',
        'share expectations no reason',
        'share expectations other reason',
        'materials needed available',
        'attend lesson',
        'child interact teacher',
        'child interact friends',
        'child clear responses',
        'child ask questions',
        'child acquire competency',
        'child show improvement',
        'child expected work independently',
        'work independently evaluation',
        'complete printed package',
        'sessions participated',
        'not participating reason',
        'E recharge card provided',
        'action to taken',
        'action to taken specify',
        'child needs pss',
        'child cant access resources',
        'homework after lesson',
        'parents supporting student',
        'completed tasks',
        'meet objectives',
        'meet objectives verified',
        'objectives verified specify',
        'additional notes'
    ]
    fc_type = [
        'pre-arabic',
        'post-arabic',
        'pre-math',
        'post-math',
        'pre-language',
        'post-language'
    ]
    for subject_type in range(len(fc_type)):
        for col_num in range(len(columns_fc)):
            wsFC.write(row_num_header_fc, col_num + len(columns_fc_student) + (len(columns_fc) * (subject_type))
                       , columns_fc[col_num], font_style)

    font_style = xlwt.XFStyle()

    rows_fc = queryset_fc.values_list(
        'enrollment_id',
        'enrollment__student__first_name',
        'enrollment__student__father_name',
        'enrollment__student__last_name',
        'fc_type',
        'facilitator_name',
        'subject_taught',
        'date_of_monitoring',
        'targeted_competencies',
        'activities_reported',
        'activities_reported_other',
        'share_expectations',
        'share_expectations_no_reason',
        'share_expectations_other_reason',
        'materials_needed_available',
        'attend_lesson',
        'child_interact_teacher',
        'child_interact_friends',
        'child_clear_responses',
        'child_ask_questions',
        'child_acquire_competency',
        'child_show_improvement',
        'child_expected_work_independently',
        'work_independently_evaluation',
        'complete_printed_package',
        'sessions_participated',
        'not_participating_reason',
        'e_recharge_card_provided',
        'action_to_taken',
        'action_to_taken_specify',
        'child_needs_pss',
        'child_cant_access_resources',
        'homework_after_lesson',
        'parents_supporting_student',
        'completed_tasks',
        'meet_objectives',
        'meet_objectives_verified',
        'objectives_verified_specify',
        'additional_notes'

    )
    enrollment_id = 0

    for row in rows_fc:
        row_num_fc += 0
        enrol_id = row[0]

        if enrollment_id != enrol_id:
            row_num_fc += 1
            enrollment_id = enrol_id
            for col_num in range(len(columns_fc_student)):
                wsFC.write(row_num_fc, col_num, row[col_num], font_style)

            row_type = row[4]
            subject_type_index = fc_type.index(row_type)

            for col_num in range(len(columns_fc)):
                wsFC.write(row_num_fc, col_num + len(columns_fc_student) + (len(columns_fc) * (subject_type_index))
                           , row[col_num + len(columns_fc_student)], font_style)

    wb_student.save(buffer)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    response = FileResponse(buffer, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="BLN.xls"'

    return response


class MemorySavingQuerysetIterator(object):

    def __init__(self, queryset, max_obj_num=1000):
        self._base_queryset = queryset
        self._generator = self._setup()
        self.max_obj_num = max_obj_num

    def _setup(self):
        for i in xrange(0, self._base_queryset.count(), self.max_obj_num):
            # By making a copy of the queryset and using that to actually access
            # the objects we ensure that there are only `max_obj_num` objects in
            # memory at any given time
            smaller_queryset = copy.deepcopy(self._base_queryset)[i:i + self.max_obj_num]
            # logger.debug('Grabbing next %s objects from DB' % self.max_obj_num)
            for obj in smaller_queryset.iterator():
                yield obj

    def __iter__(self):
        return self

    def next(self):
        return self._generator.next()


def get_outreach_child(outreach_id):
    initial = {}
    instance = OutreachChild.objects.get(id=outreach_id)
    initial['child_outreach'] = instance.id
    initial['student_first_name'] = instance.first_name
    initial['student_father_name'] = instance.outreach_caregiver.father_name
    initial['student_last_name'] = instance.outreach_caregiver.last_name
    initial['student_mother_fullname'] = instance.outreach_caregiver.mother_full_name
    initial['student_birthday_year'] = instance.birthday_year
    initial['student_birthday_month'] = instance.birthday_month
    initial['student_birthday_day'] = instance.birthday_day
    initial['student_sex'] = instance.gender
    nationality = instance.nationality

   # TODO check nationality'palestinian'
    # 3: for "فلسطينية  - من سوريا"
    # 4: for "فلسطنية - من لبنان"
    if nationality == 'syrian':
        initial['student_nationality'] = 1
    elif nationality == 'lebanese':
        initial['student_nationality'] = 5
    # elif nationality == 'palestinian':
    #     initial['student_nationality'] = 4
    elif nationality == 'iraqi':
        initial['student_nationality'] = 2
    elif nationality == 'stateless':
        initial['student_nationality'] = 7
    elif nationality == 'other':
        initial['student_nationality'] = 6
    initial['other_nationality'] = instance.nationality_other
    initial['student_address'] = instance.outreach_caregiver.address

    disability = instance.disability
    if disability == 'no':
        initial['disability'] = 1
    elif disability == 'difficulty_seeing':
        initial['disability'] = 6
    elif disability == 'difficulty_interacting_with_others':
        initial['disability'] = 9
    elif disability == 'difficulty_speaking':
        initial['disability'] = 5
    elif disability == 'intellectual_disability':
        initial['disability'] = 10
    elif disability == 'difficulty_hearing':
        initial['disability'] = 3
    elif disability == 'learning_difficulties':
        initial['disability'] = 8
    elif disability == 'difficulty_walking_or_moving_hands':
        initial['disability'] = 4

    initial['student_family_status'] = instance.family_status

    # TODO check nationality'palestinian'
    # 3: for "فلسطينية  - من سوريا"
    # 4: for "فلسطنية - من لبنان"
    main_caregiver_nationality = instance.outreach_caregiver.caregiver_nationality
    if main_caregiver_nationality == 'syrian':
        initial['main_caregiver_nationality'] = 1
    elif main_caregiver_nationality == 'lebanese':
    #     initial['main_caregiver_nationality'] = 5
    # elif main_caregiver_nationality == 'palestinian':
        initial['main_caregiver_nationality'] = 4
    elif main_caregiver_nationality == 'iraqi':
        initial['main_caregiver_nationality'] = 2
    elif main_caregiver_nationality == 'stateless':
        initial['main_caregiver_nationality'] = 7
    elif main_caregiver_nationality == 'other':
        initial['main_caregiver_nationality'] = 6
    initial['main_caregiver_nationality_other'] = instance.outreach_caregiver.caregiver_nationality_other

    initial['have_labour_single_selection'] = instance.working_status
    labour_type = instance.work_type
    if labour_type == 'manufacturing_producing':
        initial['labour_type'] = 'manufacturing'
    elif labour_type == 'garage_mechanics_workshop':
        initial['labour_type'] = ''
    elif labour_type == 'construction_site':
        initial['labour_type'] = 'building'
    elif labour_type == 'shop_restaurant_bakery_barber':
        initial['labour_type'] = 'retail_store'
    elif labour_type == 'street_connected_work__begging__vending_':
        initial['labour_type'] = 'begging'
    elif labour_type == 'agriculture_animal_herding':
        initial['labour_type'] = 'agriculture'
    elif labour_type == 'others':
        initial['labour_type'] = 'other_many_other'
    else:
        initial['labour_type'] = ''

    initial['labours_other_specify'] = instance.work_type_other

    initial['phone_number'] = instance.outreach_caregiver.primary_phone
    initial['phone_number_confirm'] = instance.outreach_caregiver.primary_phone
    initial['second_phone_number'] = instance.outreach_caregiver.secondary_phone
    initial['second_phone_number_confirm'] = instance.outreach_caregiver.secondary_phone

    main_caregiver = instance.outreach_caregiver.main_caregiver
    if main_caregiver == u'الاب':
        initial['main_caregiver'] = 'father'
        initial['caretaker_first_name'] = instance.outreach_caregiver.father_name
        initial['caretaker_last_name'] = instance.outreach_caregiver.last_name
    else:
        if main_caregiver == u'الام':
            initial['main_caregiver'] = 'mother'
        elif main_caregiver == u'اخر':
            initial['main_caregiver'] = 'other'
        initial['caretaker_first_name'] = instance.outreach_caregiver.caregiver_first_name
        initial['caretaker_last_name'] = instance.outreach_caregiver.caregiver_last_name

    initial['caretaker_middle_name'] = instance.outreach_caregiver.caregiver_father_name
    initial['caretaker_mother_name'] = instance.outreach_caregiver.caregiver_mother_name

    id_type = instance.outreach_caregiver.id_type
    if id_type == 'unhcr_registered':
        initial['id_type'] = 'UNHCR Registered'
        initial['case_number'] = instance.outreach_caregiver.unhcr_case_number
        initial['case_number_confirm'] = instance.outreach_caregiver.unhcr_case_number
        initial['parent_individual_case_number'] = instance.outreach_caregiver.caregiver_unhcr_id
        initial['parent_individual_case_number_confirm'] = instance.outreach_caregiver.caregiver_unhcr_id
        initial['individual_case_number'] = instance.child_unhcr_number
        initial['individual_case_number_confirm'] = instance.child_unhcr_number
    elif id_type == 'unhcr_recorded':
        initial['id_type'] = 'UNHCR Recorded'
        initial['recorded_number'] = instance.outreach_caregiver.unhcr_barcode
        initial['recorded_number_confirm'] = instance.outreach_caregiver.unhcr_barcode
    elif id_type == 'syrian_id':
        initial['id_type'] = 'Syrian national ID'
        initial['parent_syrian_national_number'] = instance.outreach_caregiver.caregiver_personal_id
        initial['parent_syrian_national_number_confirm'] = instance.outreach_caregiver.caregiver_personal_id
        initial['syrian_national_number'] = instance.child_personal_id
        initial['syrian_national_number_confirm'] = instance.child_personal_id
    elif id_type == 'palestinian_id':
        initial['id_type'] = 'Palestinian national ID'
        initial['sop_parent_national_number'] = instance.outreach_caregiver.caregiver_personal_id
        initial['sop_parent_national_number_confirm'] = instance.outreach_caregiver.caregiver_personal_id
        initial['sop_national_number'] = instance.child_personal_id
        initial['sop_national_number_confirm'] = instance.child_personal_id
    elif id_type == 'lebanese_id':
        initial['id_type'] = 'Lebanese national ID'
        initial['parent_national_number'] = instance.outreach_caregiver.caregiver_personal_id
        initial['parent_national_number_confirm'] = instance.outreach_caregiver.caregiver_personal_id
        initial['national_number'] = instance.child_personal_id
        initial['national_number_confirm'] = instance.child_personal_id

    return initial
