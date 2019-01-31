
from student_registration.taskapp.celery import app

import json
import httplib
import datetime
from time import mktime
from django.core.serializers.json import DjangoJSONEncoder
from openpyxl import load_workbook


@app.task
def push_household_data(base_url, token, protocol='HTTPS'):

    wb = load_workbook(filename='BTS_Outreach_form_16102017.xlsx', read_only=True)
    ws = wb['Household']

    try:
        for row in ws.rows:
            if row[0].value == 'FORMID':
                continue
            data = {}
            data['form_id'] = row[0].value
            data['partner_name'] = row[1].value if row[1].value else 'None'
            data['governorate'] = row[2].value if row[2].value else 'None'
            data['district'] = row[3].value if row[3].value else 'None'
            data['village'] = row[4].value if row[4].value else 'None'
            data['name'] = row[5].value if row[5].value else 'None'
            data['phone_number'] = row[6].value if row[6].value else '000000'
            data['residence_type'] = row[7].value if row[7].value else 'None'
            data['p_code'] = row[8].value if row[8].value else 'None'
            data['address'] = row[9].value if row[9].value else 'None'
            data['number_of_children'] = row[10].value if row[10].value else '0'
            data['barcode_number'] = row[11].value if row[11].value else '0'

            post_data(protocol=protocol, url=base_url, apifunc='/api/household/', token=token, data=data)
    except Exception as ex:
        print("---------------")
        print("error: ", ex.message)
        print(json.dumps(data, cls=DjangoJSONEncoder))
        print("---------------")
        pass


@app.task
def push_children_data(base_url, token, protocol='HTTPS'):

    # NATIONALITIES = {
    #     'Syria': 1,
    #     'Iraq': 2,
    #     'Lebanon': 5,
    #     'Other': 6,
    #     'Palestine': 4
    # }
    ID_TYPES = {
        'None': 6,
        'UNHCR': 1,
        'Lebanese': 5,
        'Other': 5,
        'Syria': 3,
    }

    wb = load_workbook(filename='BTS_Outreach_data_05022018.xlsx', read_only=True)
    ws = wb['Individuals']

    try:
        for row in ws.rows:
            if row[0].value == 'FORMID':
                continue

            data = {}
            data['form_id'] = row[0].value
            data['first_name'] = row[1].value if row[1].value else 'None'
            data['father_name'] = row[2].value if row[2].value else 'None'
            data['last_name'] = row[3].value if row[3].value else 'None'
            data['mother_fullname'] = row[4].value if row[4].value else 'None'

            # data['mother_nationality'] = NATIONALITIES[row[5].value] if row[5].value in NATIONALITIES else 6
            data['mother_nationality'] = row[5].value if row[5].value else 6

            data['birthday_year'] = row[6].value if row[6].value else '1990'
            data['birthday_month'] = row[7].value if row[7].value else '1'
            data['birthday_day'] = row[8].value if row[8].value else '1'
            data['nationality'] = row[9].value if row[9].value else 6

            data['id_type'] = ID_TYPES[row[10].value] if row[10].value in ID_TYPES else 6

            data['id_number'] = row[11].value if row[11].value else '000000'  # ID_Num
            if row[13].value:  # Case_Individual_Num
                data['id_number'] = row[13].value
            elif row[14].value:  # ID_Individual_Num
                data['id_number'] = row[14].value
            elif row[12].value:  # UNHCR_Case_Num
                data['id_number'] = row[12].value

            data['sex'] = row[15].value if row[15].value else 'Male'

            post_data(protocol=protocol, url=base_url, apifunc='/api/child/', token=token, data=data)
    except Exception as ex:
        print("---------------")
        print("error: ", ex.message)
        print(json.dumps(data, cls=DjangoJSONEncoder))
        print("---------------")
        pass


@app.task
def link_household_to_children():
    from .models import HouseHold, Child

    households = HouseHold.objects.all()
    for hh in households:
        ctr = 0
        children = Child.objects.filter(form_id=hh.form_id)
        for child in children:
            ctr += 1
            child.barcode_subset = '{}-{}'.format(hh.barcode_number, ctr)
            child.household = hh
            print(child.id)
            child.save()


class MyEncoder(json.JSONEncoder):

    def default(self, obj):
        if isinstance(obj, datetime.datetime) or isinstance(obj, datetime.date):
            return int(mktime(obj.timetuple()))

        return json.JSONEncoder.default(self, obj)

    def decode(self, obj):
        if isinstance(obj, datetime.datetime):
            return int(mktime(obj.timetuple()))

        return json.JSONEncoder.default(self, obj)


def post_data(protocol, url, apifunc, token, data):

    params = json.dumps(data, cls=MyEncoder)

    headers = {"Content-type": "application/json", "Authorization": token, "HTTP_REFERER": url, "Cookie": "token="+token}

    # if protocol == 'HTTPS':
    #     conn = httplib.HTTPSConnection(url)
    # else:
    #     conn = httplib.HTTPConnection(url)
    # conn.request('POST', apifunc, params, headers)
    # response = conn.getresponse()
    # result = response.read()
    #
    # if not response.status == 201:
    #     if response.status == 400:
    #         raise Exception(str(response.status) + response.reason + response.read())
    #     else:
    #         raise Exception(str(response.status) + response.reason)
    #
    # conn.close()
    #
    # return result
